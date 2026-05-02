import os
import sys

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


def _obtener_carpeta_base():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


CARPETA_BASE = _obtener_carpeta_base()
ARCHIVO_EXCEL = os.path.join(CARPETA_BASE, "fashion_reset.xlsx")


def obtener_ruta_en_base(*partes):
    return os.path.join(CARPETA_BASE, *partes)

# FUNCION AUXILIAR PARA LIMPIAR TEXTOS
def limpiar_texto(texto):
    return " ".join(texto.strip().upper().split())

# FUNCION AUXILIAR PARA LEER FECHAS
def convertir_fecha(texto_fecha):
    if isinstance(texto_fecha, datetime):
        return texto_fecha

    texto_fecha = str(texto_fecha).strip()

    for formato in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto_fecha, formato)
        except ValueError:
            pass

    raise ValueError("FECHA INVALIDA")


def normalizar_obs_descuento_proveedora(tipo_pago, obs_venta):
    tipo_pago = limpiar_texto(str(tipo_pago))
    obs_venta = limpiar_texto(str(obs_venta))

    if tipo_pago != "DESCUENTO A PROVEEDORA":
        return obs_venta

    if not obs_venta:
        return ""

    if obs_venta.startswith("DESCUENTO_PROVEEDORA:"):
        return obs_venta

    return f"DESCUENTO_PROVEEDORA: {obs_venta}"


def _asegurar_columnas_remarque(ws_ingresos):
    encabezados_requeridos = ["PRECIO_REMARCADO", "FECHA_REMARQUE", "DECISION_PROVEEDORA"]
    encabezados_actuales = {}

    for columna in range(1, ws_ingresos.max_column + 1):
        nombre = str(ws_ingresos.cell(row=1, column=columna).value or "").strip().upper()
        if nombre:
            encabezados_actuales[nombre] = columna

    for encabezado in encabezados_requeridos:
        if encabezado not in encabezados_actuales:
            nueva_columna = ws_ingresos.max_column + 1
            ws_ingresos.cell(row=1, column=nueva_columna).value = encabezado
            encabezados_actuales[encabezado] = nueva_columna

    return encabezados_actuales

                                                
                                                
                                                
                                                #CARGA DE INGRESO
# ABRIR ARCHIVO EXCEL Y HOJA INGRESOS
def _guardar_ingreso_en_excel(
    fecha_ingreso,
    codigo_proveedora,
    numero_prenda,
    articulo,
    marca,
    talle,
    color,
    precio_texto,
    obs_ingreso,
    precio_obligatorio=False
):
    fecha_ingreso = str(fecha_ingreso).strip()
    codigo_proveedora = limpiar_texto(str(codigo_proveedora))
    numero_prenda = limpiar_texto(str(numero_prenda))
    articulo = limpiar_texto(str(articulo))
    marca = limpiar_texto(str(marca))
    talle = limpiar_texto(str(talle))
    color = limpiar_texto(str(color))
    obs_ingreso = limpiar_texto(str(obs_ingreso))
    precio_texto = str(precio_texto).strip()

    if not codigo_proveedora:
        return False, "ERROR: CODIGO_PROVEEDORA es obligatorio."

    if not numero_prenda or not articulo:
        return False, "ERROR: NUMERO_PRENDA y ARTICULO son obligatorios."

    codigo_prenda = f"{codigo_proveedora}{numero_prenda}"

    try:
        datetime.strptime(fecha_ingreso, "%d/%m/%Y")
    except ValueError:
        return False, "ERROR: FECHA_INGRESO invalida. Usa el formato DD/MM/YYYY."

    precio = ""
    if precio_texto:
        try:
            precio = int(precio_texto)
        except ValueError:
            return False, "ERROR: PRECIO invalido. Debe ser un numero entero."
    elif precio_obligatorio:
        return False, "ERROR: PRECIO invalido. Debe ser un numero entero."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    codigos_existentes = set()
    for fila in ws.iter_rows(min_row=2, values_only=True):
        codigo_existente = fila[2]
        if codigo_existente:
            codigos_existentes.add(str(codigo_existente).strip().upper())

    if codigo_prenda in codigos_existentes:
        return False, f"ERROR: El codigo completo {codigo_prenda} ya existe en INGRESOS."

    ws.append([
        fecha_ingreso,
        codigo_proveedora,
        codigo_prenda,
        articulo,
        marca,
        talle,
        color,
        precio,
        "DISPONIBLE",
        "",
        "",
        "",
        "",
        "",
        obs_ingreso,
        ""
    ])

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, codigo_prenda


def guardar_ingreso_desde_gui(fecha_ingreso, codigo_proveedora, numero_prenda, articulo, marca, talle, color, precio_texto, obs_ingreso):
    return _guardar_ingreso_en_excel(
        fecha_ingreso,
        codigo_proveedora,
        numero_prenda,
        articulo,
        marca,
        talle,
        color,
        precio_texto,
        obs_ingreso,
        precio_obligatorio=False
    )


def buscar_ingreso_por_codigo(codigo_prenda):
    codigo_prenda = str(codigo_prenda).strip().upper()
    if not codigo_prenda:
        return False, "CODIGO VACIO"

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    for fila in ws.iter_rows(min_row=2, values_only=True):
        valor_codigo = fila[2]
        if valor_codigo and str(valor_codigo).strip().upper() == codigo_prenda:
            estado = str(fila[8] or "").strip().upper()
            if estado != "DISPONIBLE":
                return False, "NO DISPONIBLE"

            return True, {
                "articulo": str(fila[3] or "").strip(),
                "marca": str(fila[4] or "").strip(),
                "talle": str(fila[5] or "").strip(),
                "color": str(fila[6] or "").strip(),
                "precio_lista": str(fila[7] or "").strip()
            }

    return False, "NO ENCONTRADO"


def guardar_venta_desde_gui(
    fecha_venta,
    cliente,
    tipo_pago,
    validacion,
    codigo_prenda,
    articulo,
    marca,
    talle,
    color,
    precio_lista,
    precio_venta_texto,
    obs_venta
):
    fecha_venta = str(fecha_venta).strip()
    cliente = limpiar_texto(str(cliente))
    tipo_pago = limpiar_texto(str(tipo_pago))
    validacion = limpiar_texto(str(validacion))
    codigo_prenda = str(codigo_prenda).strip().upper()
    articulo = str(articulo).strip()
    marca = str(marca).strip()
    talle = str(talle).strip()
    color = str(color).strip()
    precio_lista = str(precio_lista).strip()
    precio_venta_texto = str(precio_venta_texto).strip()
    obs_venta = limpiar_texto(str(obs_venta))

    if not fecha_venta or not cliente or not tipo_pago or not validacion or not codigo_prenda:
        return False, "ERROR: Fecha, cliente, tipo pago, validación y código prenda son obligatorios."

    if tipo_pago == "DESCUENTO A PROVEEDORA":
        if not obs_venta:
            return False, "ERROR: Para DESCUENTO A PROVEEDORA, debes indicar el codigo de la proveedora."

    try:
        datetime.strptime(fecha_venta, "%d/%m/%Y")
    except ValueError:
        return False, "ERROR: FECHA_VENTA invalida. Usa el formato DD/MM/YYYY."

    if not precio_venta_texto:
        precio_venta_texto = precio_lista

    try:
        precio_venta = int(str(precio_venta_texto).strip())
    except ValueError:
        return False, "ERROR: PRECIO_VENTA invalido. Debe ser un numero entero."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS o VENTAS en el archivo Excel."

    fila_encontrada = None
    for fila in ws_ingresos.iter_rows(min_row=2):
        valor_codigo = fila[2].value
        if valor_codigo and str(valor_codigo).strip().upper() == codigo_prenda:
            fila_encontrada = fila
            break

    if not fila_encontrada:
        return False, f"ERROR: No se encontro la prenda {codigo_prenda} en INGRESOS."

    estado = str(fila_encontrada[8].value or "").strip().upper()
    if estado != "DISPONIBLE":
        return False, f"ERROR: La prenda {codigo_prenda} no esta DISPONIBLE."

    fila_numero = fila_encontrada[0].row
    ws_ingresos.cell(row=fila_numero, column=9, value="VENDIDO")
    ws_ingresos.cell(row=fila_numero, column=10, value=fecha_venta)
    ws_ingresos.cell(row=fila_numero, column=11, value=precio_venta)
    ws_ingresos.cell(row=fila_numero, column=12, value=cliente)
    ws_ingresos.cell(row=fila_numero, column=13, value=tipo_pago)
    ws_ingresos.cell(row=fila_numero, column=14, value=validacion)
    ws_ingresos.cell(row=fila_numero, column=16, value=obs_venta)

    ws_ventas.append([
        fecha_venta,
        fila_encontrada[1].value,
        codigo_prenda,
        articulo,
        marca,
        talle,
        color,
        precio_lista,
        precio_venta,
        cliente,
        tipo_pago,
        validacion,
        obs_venta
    ])

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, codigo_prenda


def cargar_ingreso(fecha_ingreso, codigo_proveedora):
    print("\n=== CARGA DE INGRESO ===")

 # ENTRADA DE DATOS
    numero_prenda = limpiar_texto(input("NUMERO_PRENDA: "))
    articulo = limpiar_texto(input("ARTICULO: "))
    marca = limpiar_texto(input("MARCA: "))
    talle = limpiar_texto(input("TALLE: "))
    color = limpiar_texto(input("COLOR: "))
    precio_texto = input("PRECIO: ").strip()
    obs_ingreso = limpiar_texto(input("OBS_INGRESO: "))

    resultado, mensaje = _guardar_ingreso_en_excel(
        fecha_ingreso,
        codigo_proveedora,
        numero_prenda,
        articulo,
        marca,
        talle,
        color,
        precio_texto,
        obs_ingreso,
        precio_obligatorio=True
    )

    if not resultado:
        print(mensaje)
        return

    codigo_prenda = mensaje
    articulo = limpiar_texto(articulo)
    marca = limpiar_texto(marca)
    talle = limpiar_texto(talle)
    color = limpiar_texto(color)
    obs_ingreso = limpiar_texto(obs_ingreso)
    precio = int(str(precio_texto).strip())

    print(f"""
        CODIGO_PRENDA: {codigo_prenda}
        ARTICULO: {articulo}
        MARCA: {marca}
        TALLE: {talle}
        COLOR: {color}
        PRECIO: {precio}
        ESTADO: DISPONIBLE
        OBS_INGRESO: {obs_ingreso}
        """)




                                            # CARGA DE VENTA
def cargar_venta():
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return
    except FileNotFoundError:
        print("ERROR: No se encontro el archivo fashion_reset.xlsx.")
        return
    except KeyError:
        print("ERROR: No existe la hoja INGRESOS o VENTAS en el archivo Excel.")
        return

    print("\n=== CARGA DE VENTA ===")

    # DATOS GENERALES DE LA VENTA
    fecha_venta = input("FECHA_VENTA (DD/MM/YYYY): ").strip()
    cliente = limpiar_texto(input("CLIENTE: "))

    print("""
TIPO_PAGO:
1 - TRANSFERENCIA
2 - EFECTIVO
3 - DESCUENTO A PROVEEDORA
4 - OTRO
""")

    opcion_pago = input("ELIGE TIPO_PAGO: ").strip()

    if opcion_pago == "1":
        tipo_pago = "TRANSFERENCIA"
    elif opcion_pago == "2":
        tipo_pago = "EFECTIVO"
    elif opcion_pago == "3":
        tipo_pago = "DESCUENTO A PROVEEDORA"
    elif opcion_pago == "4":
        tipo_pago = "OTRO"
    else:
        print("ERROR: TIPO_PAGO invalido.")
        return

    print("""
VALIDACION:
1 - PENDIENTE
2 - PAGADO
""")

    opcion_validacion = input("ELIGE VALIDACION: ").strip()

    if opcion_validacion == "1":
        validacion = "PENDIENTE"
    elif opcion_validacion == "2":
        validacion = "PAGADO"
    else:
        print("ERROR: VALIDACION invalida.")
        return

    # VALIDAR FECHA_VENTA
    try:
        datetime.strptime(fecha_venta, "%d/%m/%Y")
    except ValueError:
        print("ERROR: FECHA_VENTA invalida. Usa el formato DD/MM/YYYY.")
        return

    # CARGA MULTIPLE DE PRENDAS EN LA MISMA VENTA
    while True:
        # PEDIR CODIGO DE PRENDA
        codigo_prenda = limpiar_texto(input("CODIGO_PRENDA: "))

        if not codigo_prenda:
            print("ERROR: CODIGO_PRENDA es obligatorio.")
            otra = input("¿CARGAR OTRA PRENDA EN ESTA MISMA VENTA? (S/N): ").strip().upper()
            if otra != "S":
                break
            continue

        # BUSCAR PRENDA EN INGRESOS
        fila_encontrada = None

        for fila in ws_ingresos.iter_rows(min_row=2):
            valor_codigo = fila[2].value  # COLUMNA CODIGO_PRENDA
            if valor_codigo and str(valor_codigo).strip().upper() == codigo_prenda:
                fila_encontrada = fila
                break

        if not fila_encontrada:
            print(f"ERROR: No se encontro la prenda {codigo_prenda} en INGRESOS.")
            otra = input("¿CARGAR OTRA PRENDA EN ESTA MISMA VENTA? (S/N): ").strip().upper()
            if otra != "S":
                break
            continue

        # LEER DATOS DE LA FILA ENCONTRADA
        codigo_proveedora = fila_encontrada[1].value
        articulo = fila_encontrada[3].value
        marca = fila_encontrada[4].value
        talle = fila_encontrada[5].value
        color = fila_encontrada[6].value
        precio_lista = fila_encontrada[7].value
        estado = fila_encontrada[8].value

        # VALIDAR ESTADO
        if str(estado).strip().upper() != "DISPONIBLE":
            print(f"ERROR: La prenda {codigo_prenda} no esta DISPONIBLE. Estado actual: {estado}")
            otra = input("¿CARGAR OTRA PRENDA EN ESTA MISMA VENTA? (S/N): ").strip().upper()
            if otra != "S":
                break
            continue

        # MOSTRAR RESUMEN DE LA PRENDA
        print(f"""
CODIGO_PRENDA: {codigo_prenda}
CODIGO_PROVEEDORA: {codigo_proveedora}
ARTICULO: {articulo}
MARCA: {marca}
TALLE: {talle}
COLOR: {color}
PRECIO_LISTA: {precio_lista}
ESTADO: {estado}
""")

        # DATOS DE VENTA DE ESTA PRENDA
        precio_venta_texto = input(f"PRECIO_VENTA ({precio_lista}): ").strip()
        if not precio_venta_texto:
            precio_venta_texto = str(precio_lista)

        if tipo_pago == "DESCUENTO A PROVEEDORA":
            codigo_proveedora_descuento = limpiar_texto(input("CODIGO_PROVEEDORA A DESCONTAR: "))
            if not codigo_proveedora_descuento:
                print("ERROR: CODIGO_PROVEEDORA A DESCONTAR es obligatorio.")
                otra = input("¿CARGAR OTRA PRENDA EN ESTA MISMA VENTA? (S/N): ").strip().upper()
                if otra != "S":
                    break
                continue

            obs_venta = f"DESCUENTO_PROVEEDORA: {codigo_proveedora_descuento}"
        else:
            obs_venta = limpiar_texto(input("OBS_VENTA: "))

        # VALIDAR PRECIO_VENTA
        try:
            precio_venta = int(precio_venta_texto)
        except ValueError:
            print("ERROR: PRECIO_VENTA invalido. Debe ser un numero entero.")
            otra = input("¿CARGAR OTRA PRENDA EN ESTA MISMA VENTA? (S/N): ").strip().upper()
            if otra != "S":
                break
            continue

        # ACTUALIZAR FILA EN INGRESOS
        fila_numero = fila_encontrada[0].row

        ws_ingresos.cell(row=fila_numero, column=9, value="VENDIDO")          # ESTADO
        ws_ingresos.cell(row=fila_numero, column=10, value=fecha_venta)       # FECHA_VENTA
        ws_ingresos.cell(row=fila_numero, column=11, value=precio_venta)      # PRECIO_VENTA
        ws_ingresos.cell(row=fila_numero, column=12, value=cliente)           # CLIENTE
        ws_ingresos.cell(row=fila_numero, column=13, value=tipo_pago)         # TIPO_PAGO
        ws_ingresos.cell(row=fila_numero, column=14, value=validacion)        # VALIDACION
        ws_ingresos.cell(row=fila_numero, column=16, value=obs_venta)         # OBS_VENTA

        # GUARDAR NUEVA FILA EN VENTAS
        ws_ventas.append([
            fecha_venta,
            codigo_proveedora,
            codigo_prenda,
            articulo,
            marca,
            talle,
            color,
            precio_lista,
            precio_venta,
            cliente,
            tipo_pago,
            validacion,
            obs_venta
        ])

        # GUARDAR CAMBIOS EN EL EXCEL
        try:
            wb.save(ARCHIVO_EXCEL)
            print(f"""
VENTA REGISTRADA CORRECTAMENTE
CODIGO_PRENDA: {codigo_prenda}
CLIENTE: {cliente}
PRECIO_LISTA: {precio_lista}
PRECIO_VENTA: {precio_venta}
TIPO_PAGO: {tipo_pago}
VALIDACION: {validacion}
OBS_VENTA: {obs_venta}
""")
        except PermissionError:
            print("ERROR: No se pudo guardar porque el archivo Excel esta abierto.")
            break

        otra = input("¿CARGAR OTRA PRENDA EN ESTA MISMA VENTA? (S/N): ").strip().upper()
        if otra != "S":
            break
 
 
 
def calcular_resumen_general(mes_texto, anio_texto):
    try:
        mes = int(str(mes_texto).strip())
        if mes < 1 or mes > 12:
            return False, "ERROR: El MES debe estar entre 1 y 12."
    except ValueError:
        return False, "ERROR: MES invalido."

    try:
        anio = int(str(anio_texto).strip())
    except ValueError:
        return False, "ERROR: AÑO invalido."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja VENTAS en el archivo Excel."

    total_prendas = 0
    total_vendido = 0
    total_descuentos = 0
    resumen_proveedoras = {}

    for fila in ws_ventas.iter_rows(min_row=2, values_only=True):
        fecha_venta = fila[0]
        codigo_proveedora = str(fila[1] or "").strip().upper()
        precio_venta = fila[8]
        tipo_pago = fila[10]
        validacion = fila[11]

        try:
            fecha_obj = convertir_fecha(fecha_venta)
        except Exception:
            continue

        if fecha_obj.month == mes and fecha_obj.year == anio and str(validacion).strip().upper() == "PAGADO":
            if not isinstance(precio_venta, (int, float)):
                continue

            precio_venta = int(precio_venta)
            total_prendas += 1
            total_vendido += precio_venta

            if str(tipo_pago).strip().upper() == "DESCUENTO A PROVEEDORA":
                total_descuentos += precio_venta

            if codigo_proveedora not in resumen_proveedoras:
                resumen_proveedoras[codigo_proveedora] = {
                    "codigo_proveedora": codigo_proveedora,
                    "cantidad_prendas": 0,
                    "total_vendido": 0,
                    "total_descuentos": 0
                }

            resumen_proveedoras[codigo_proveedora]["cantidad_prendas"] += 1
            resumen_proveedoras[codigo_proveedora]["total_vendido"] += precio_venta

            if str(tipo_pago).strip().upper() == "DESCUENTO A PROVEEDORA":
                resumen_proveedoras[codigo_proveedora]["total_descuentos"] += precio_venta

    if total_prendas == 0:
        return False, "NO SE ENCONTRARON VENTAS PAGADAS EN ESE PERIODO."

    total_proveedoras = int(total_vendido * 0.60)
    total_fashion_reset = int(total_vendido * 0.40)
    total_neto_a_pagar = int(total_proveedoras - total_descuentos)

    detalle_proveedoras = []
    for codigo_proveedora in sorted(resumen_proveedoras.keys()):
        datos = resumen_proveedoras[codigo_proveedora]
        comision_proveedora = int(datos["total_vendido"] * 0.60)
        saldo_final = int(comision_proveedora - datos["total_descuentos"])
        detalle_proveedoras.append({
            "codigo_proveedora": codigo_proveedora,
            "cantidad_prendas": datos["cantidad_prendas"],
            "total_vendido": datos["total_vendido"],
            "comision_proveedora": comision_proveedora,
            "descuentos": datos["total_descuentos"],
            "saldo_final": saldo_final
        })

    return True, {
        "mes": mes,
        "anio": anio,
        "cantidad_total_vendida": total_prendas,
        "total_vendido": total_vendido,
        "total_proveedoras": total_proveedoras,
        "total_fashion_reset": total_fashion_reset,
        "total_descuentos": total_descuentos,
        "total_neto_a_pagar": total_neto_a_pagar,
        "detalle_proveedoras": detalle_proveedoras
    }


                                    # RESUMEN GENERAL DEL MES
def resumen_general_mes():
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return
    except FileNotFoundError:
        print("ERROR: No se encontro el archivo fashion_reset.xlsx.")
        return
    except KeyError:
        print("ERROR: No existe la hoja VENTAS en el archivo Excel.")
        return

    print("\n=== RESUMEN GENERAL DEL MES ===")

    # PEDIR DATOS DE BUSQUEDA
    mes_texto = input("MES (1-12): ").strip()
    anio_texto = input("AÑO: ").strip()

    # VALIDAR MES
    try:
        mes = int(mes_texto)
        if mes < 1 or mes > 12:
            print("ERROR: El MES debe estar entre 1 y 12.")
            return
    except ValueError:
        print("ERROR: MES invalido.")
        return

    # VALIDAR AÑO
    try:
        anio = int(anio_texto)
    except ValueError:
        print("ERROR: AÑO invalido.")
        return

    print(f"""
BUSCANDO RESUMEN GENERAL...
MES: {mes}
AÑO: {anio}
""")

    # VARIABLES GENERALES
    total_prendas = 0
    total_vendido = 0
    total_descuentos = 0

    # ACUMULADOR POR PROVEEDORA
    resumen_proveedoras = {}

    # RECORRER HOJA VENTAS
    for fila in ws_ventas.iter_rows(min_row=2, values_only=True):
        fecha_venta = fila[0]
        codigo_proveedora = str(fila[1]).strip().upper()
        precio_venta = fila[8]
        tipo_pago = fila[10]
        validacion = fila[11]

        # VALIDAR FECHA
        try:
            fecha_obj = convertir_fecha(fecha_venta)
        except:
            continue

        # FILTRAR MES, AÑO Y SOLO PAGADO
        if fecha_obj.month == mes and fecha_obj.year == anio and str(validacion).strip().upper() == "PAGADO":
            total_prendas += 1
            total_vendido += precio_venta

            if str(tipo_pago).strip().upper() == "DESCUENTO A PROVEEDORA":
                total_descuentos += precio_venta

            # CREAR PROVEEDORA SI NO EXISTE
            if codigo_proveedora not in resumen_proveedoras:
                resumen_proveedoras[codigo_proveedora] = {
                    "cantidad_prendas": 0,
                    "total_vendido": 0,
                    "total_descuentos": 0
                }

            # ACUMULAR POR PROVEEDORA
            resumen_proveedoras[codigo_proveedora]["cantidad_prendas"] += 1
            resumen_proveedoras[codigo_proveedora]["total_vendido"] += precio_venta

            if str(tipo_pago).strip().upper() == "DESCUENTO A PROVEEDORA":
                resumen_proveedoras[codigo_proveedora]["total_descuentos"] += precio_venta

    # VALIDAR SI NO HAY VENTAS
    if total_prendas == 0:
        print("NO SE ENCONTRARON VENTAS PAGADAS EN ESE PERIODO.")
        return

    # CALCULOS GENERALES
    total_proveedoras = int(total_vendido * 0.60)
    total_fashion_reset = int(total_vendido * 0.40)
    total_neto_a_pagar = int(total_proveedoras - total_descuentos)

    # MOSTRAR RESUMEN GENERAL
    print(f"""
=== RESUMEN GENERAL DEL MES ===
CANTIDAD_TOTAL_VENDIDA: {total_prendas}
TOTAL_VENDIDO: {total_vendido}
TOTAL_PROVEEDORAS (60%): {total_proveedoras}
TOTAL_FASHION_RESET (40%): {total_fashion_reset}
TOTAL_DESCUENTOS_A_PROVEEDORAS: {total_descuentos}
TOTAL_NETO_A_PAGAR: {total_neto_a_pagar}
""")

    # MOSTRAR DESGLOSE POR PROVEEDORA
    print("=== DESGLOSE POR PROVEEDORA ===")

    for codigo_proveedora, datos in resumen_proveedoras.items():
        comision_proveedora = int(datos["total_vendido"] * 0.60)
        saldo_final = int(comision_proveedora - datos["total_descuentos"])

        print(f"""
CODIGO_PROVEEDORA: {codigo_proveedora}
CANTIDAD_PRENDAS: {datos["cantidad_prendas"]}
TOTAL_VENDIDO: {datos["total_vendido"]}
COMISION_PROVEEDORA (60%): {comision_proveedora}
DESCUENTOS: {datos["total_descuentos"]}
SALDO_FINAL_A_PAGAR: {saldo_final}
""")



def calcular_ventas_pendientes():
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja VENTAS en el archivo Excel."

    cantidad_pendientes = 0
    total_pendiente = 0
    ventas_pendientes = []

    for fila in ws_ventas.iter_rows(min_row=2, values_only=True):
        fecha_venta = fila[0]
        codigo_proveedora = fila[1]
        codigo_prenda = fila[2]
        articulo = fila[3]
        precio_venta = fila[8]
        cliente = fila[9]
        tipo_pago = fila[10]
        validacion = str(fila[11] or "").strip().upper()

        if validacion == "PENDIENTE":
            cantidad_pendientes += 1
            if isinstance(precio_venta, (int, float)):
                total_pendiente += int(precio_venta)

            ventas_pendientes.append({
                "fecha_venta": fecha_venta,
                "codigo_proveedora": str(codigo_proveedora or "").strip().upper(),
                "codigo_prenda": str(codigo_prenda or "").strip().upper(),
                "articulo": str(articulo or "").strip(),
                "precio_venta": int(precio_venta) if isinstance(precio_venta, (int, float)) else 0,
                "cliente": str(cliente or "").strip(),
                "tipo_pago": str(tipo_pago or "").strip(),
                "validacion": validacion
            })

    if cantidad_pendientes == 0:
        return False, "NO HAY VENTAS PENDIENTES DE VALIDACION."

    return True, {
        "cantidad_pendientes": cantidad_pendientes,
        "total_importe_pendiente": total_pendiente,
        "ventas_pendientes": ventas_pendientes
    }


def calcular_prendas_vencidas():
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    hoy = datetime.now().date()
    prendas_vencidas = []

    for fila in ws_ingresos.iter_rows(min_row=2, values_only=True):
        estado = str(fila[8] or "").strip().upper()
        if estado != "DISPONIBLE":
            continue

        try:
            fecha_convertida = convertir_fecha(fila[0])
        except ValueError:
            continue

        dias_en_inventario = (hoy - fecha_convertida.date()).days
        if dias_en_inventario < 90:
            continue

        prendas_vencidas.append({
            "fecha_ingreso": fecha_convertida,
            "codigo_proveedora": str(fila[1] or "").strip(),
            "codigo_prenda": str(fila[2] or "").strip(),
            "articulo": str(fila[3] or "").strip(),
            "marca": str(fila[4] or "").strip(),
            "talle": str(fila[5] or "").strip(),
            "color": str(fila[6] or "").strip(),
            "precio": fila[7] if isinstance(fila[7], (int, float)) else str(fila[7] or "").strip(),
            "estado": str(fila[8] or "").strip(),
            "dias_en_inventario": dias_en_inventario
        })

    prendas_vencidas.sort(
        key=lambda fila: (-fila["dias_en_inventario"], fila["codigo_prenda"])
    )

    if not prendas_vencidas:
        return False, "NO HAY PRENDAS VENCIDAS PARA MOSTRAR."

    return True, {
        "cantidad_prendas": len(prendas_vencidas),
        "prendas_vencidas": prendas_vencidas
    }


def obtener_prenda_para_remarque(codigo_prenda):
    codigo_prenda = limpiar_texto(str(codigo_prenda))

    if not codigo_prenda:
        return False, "ERROR: CODIGO_PRENDA es obligatorio."

    resultado, datos = obtener_prenda_desde_gui(codigo_prenda)
    if not resultado:
        return False, datos

    if str(datos["estado"] or "").strip().upper() != "DISPONIBLE":
        return False, "ERROR: SOLO SE PUEDEN REMARCAR PRENDAS DISPONIBLES."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    encabezados = _asegurar_columnas_remarque(ws_ingresos)

    fila_prenda = None
    for fila in ws_ingresos.iter_rows(min_row=2):
        if str(fila[2].value or "").strip().upper() == codigo_prenda:
            fila_prenda = fila
            break

    if fila_prenda is None:
        return False, "ERROR: NO SE ENCONTRO LA PRENDA."

    try:
        fecha_ingreso = convertir_fecha(fila_prenda[0].value)
    except ValueError:
        return False, "ERROR: LA FECHA_INGRESO DE LA PRENDA ES INVALIDA."

    dias_en_inventario = (datetime.now().date() - fecha_ingreso.date()).days
    if dias_en_inventario < 90:
        return False, "ERROR: LA PRENDA TODAVIA NO CUMPLE 3 MESES EN INVENTARIO."

    col_precio_remarcado = encabezados["PRECIO_REMARCADO"] - 1
    col_fecha_remarque = encabezados["FECHA_REMARQUE"] - 1
    col_decision = encabezados["DECISION_PROVEEDORA"] - 1

    datos.update({
        "fecha_ingreso": fecha_ingreso,
        "dias_en_inventario": dias_en_inventario,
        "precio_remarcado": fila_prenda[col_precio_remarcado].value,
        "fecha_remarque": fila_prenda[col_fecha_remarque].value,
        "decision_proveedora": str(fila_prenda[col_decision].value or "").strip(),
    })
    return True, datos


def registrar_remarque_desde_gui(codigo_prenda, precio_remarcado_texto, fecha_remarque=None):
    codigo_prenda = limpiar_texto(str(codigo_prenda))
    precio_remarcado_texto = str(precio_remarcado_texto).strip()

    if not codigo_prenda:
        return False, "ERROR: CODIGO_PRENDA es obligatorio."

    try:
        precio_remarcado = int(precio_remarcado_texto.replace(".", "").replace(",", ""))
    except ValueError:
        return False, "ERROR: PRECIO_REMARCADO invalido. Debe ser un numero entero."

    resultado, datos = obtener_prenda_para_remarque(codigo_prenda)
    if not resultado:
        return False, datos

    if fecha_remarque is None:
        fecha_remarque = datetime.now().strftime("%d/%m/%Y")
    else:
        fecha_remarque = str(fecha_remarque).strip()

    try:
        datetime.strptime(fecha_remarque, "%d/%m/%Y")
    except ValueError:
        return False, "ERROR: FECHA_REMARQUE invalida. Usa el formato DD/MM/YYYY."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    encabezados = _asegurar_columnas_remarque(ws_ingresos)

    fila_prenda = None
    for fila in ws_ingresos.iter_rows(min_row=2):
        if str(fila[2].value or "").strip().upper() == codigo_prenda:
            fila_prenda = fila
            break

    if fila_prenda is None:
        return False, "ERROR: NO SE ENCONTRO LA PRENDA."

    fila_prenda[encabezados["PRECIO_REMARCADO"] - 1].value = precio_remarcado
    fila_prenda[encabezados["FECHA_REMARQUE"] - 1].value = fecha_remarque
    fila_prenda[encabezados["DECISION_PROVEEDORA"] - 1].value = "PENDIENTE"

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, {
        "codigo_prenda": datos["codigo_prenda"],
        "articulo": datos["articulo"],
        "precio_actual": datos["precio"],
        "precio_remarcado": precio_remarcado,
        "fecha_remarque": fecha_remarque,
        "decision_proveedora": "PENDIENTE",
    }


def obtener_lote_remarque_proveedora_desde_gui(codigo_proveedora):
    codigo_proveedora = limpiar_texto(str(codigo_proveedora))

    if not codigo_proveedora:
        return False, "ERROR: CODIGO_PROVEEDORA es obligatorio."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    nombre_proveedora = ""
    if "PROVEEDORAS" in wb.sheetnames:
        ws_proveedoras = wb["PROVEEDORAS"]
        for fila in ws_proveedoras.iter_rows(min_row=2, values_only=True):
            if str(fila[0] or "").strip().upper() == codigo_proveedora:
                nombre_proveedora = str(fila[1] or "").strip()
                break

    hoy = datetime.now().date()
    encabezados = _asegurar_columnas_remarque(ws_ingresos)
    prendas_remarque = []

    for fila in ws_ingresos.iter_rows(min_row=2):
        codigo_proveedora_fila = str(fila[1].value or "").strip().upper()
        estado = str(fila[8].value or "").strip().upper()

        if codigo_proveedora_fila != codigo_proveedora or estado != "DISPONIBLE":
            continue

        try:
            fecha_ingreso = convertir_fecha(fila[0].value)
        except ValueError:
            continue

        dias_en_inventario = (hoy - fecha_ingreso.date()).days
        if dias_en_inventario < 90:
            continue

        prendas_remarque.append({
            "codigo_prenda": str(fila[2].value or "").strip(),
            "articulo": str(fila[3].value or "").strip(),
            "marca": str(fila[4].value or "").strip(),
            "talle": str(fila[5].value or "").strip(),
            "color": str(fila[6].value or "").strip(),
            "precio_actual": fila[7].value if isinstance(fila[7].value, (int, float)) else str(fila[7].value or "").strip(),
            "dias_en_inventario": dias_en_inventario,
            "precio_remarcado": fila[encabezados["PRECIO_REMARCADO"] - 1].value,
            "fecha_remarque": fila[encabezados["FECHA_REMARQUE"] - 1].value,
            "decision_proveedora": str(fila[encabezados["DECISION_PROVEEDORA"] - 1].value or "").strip(),
        })

    prendas_remarque.sort(
        key=lambda fila: (-fila["dias_en_inventario"], fila["codigo_prenda"])
    )

    if not prendas_remarque:
        return False, "NO HAY PRENDAS VENCIDAS Y DISPONIBLES PARA ESTA PROVEEDORA."

    return True, {
        "codigo_proveedora": codigo_proveedora,
        "nombre_proveedora": nombre_proveedora,
        "cantidad_prendas": len(prendas_remarque),
        "prendas": prendas_remarque,
    }


def guardar_lote_remarque_desde_gui(codigo_proveedora, prendas_remarque, fecha_remarque=None):
    codigo_proveedora = limpiar_texto(str(codigo_proveedora))

    if not codigo_proveedora:
        return False, "ERROR: CODIGO_PROVEEDORA es obligatorio."

    if fecha_remarque is None:
        fecha_remarque = datetime.now().strftime("%d/%m/%Y")
    else:
        fecha_remarque = str(fecha_remarque).strip()

    try:
        datetime.strptime(fecha_remarque, "%d/%m/%Y")
    except ValueError:
        return False, "ERROR: FECHA_REMARQUE invalida. Usa el formato DD/MM/YYYY."

    filas_a_guardar = []
    codigos_duplicados = set()
    codigos_vistos = set()

    for fila in prendas_remarque:
        codigo_prenda = limpiar_texto(str(fila.get("codigo_prenda", "")))
        precio_texto = str(fila.get("precio_remarcado", "")).strip()

        if not codigo_prenda and not precio_texto:
            continue

        if not codigo_prenda:
            return False, "ERROR: HAY UNA FILA DE REMARQUE SIN CODIGO_PRENDA."

        if not precio_texto:
            continue

        if codigo_prenda in codigos_vistos:
            codigos_duplicados.add(codigo_prenda)
            continue

        try:
            precio_remarcado = int(precio_texto.replace(".", "").replace(",", ""))
        except ValueError:
            return False, f"ERROR: PRECIO_REMARCADO invalido para {codigo_prenda}."

        filas_a_guardar.append({
            "codigo_prenda": codigo_prenda,
            "precio_remarcado": precio_remarcado,
        })
        codigos_vistos.add(codigo_prenda)

    if codigos_duplicados:
        return False, f"ERROR: HAY CODIGOS_PRENDA DUPLICADOS EN EL LOTE: {', '.join(sorted(codigos_duplicados))}."

    if not filas_a_guardar:
        return False, "ERROR: No hay prendas con PRECIO_REMARCADO para guardar."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    encabezados = _asegurar_columnas_remarque(ws_ingresos)
    filas_por_codigo = {}
    for fila in ws_ingresos.iter_rows(min_row=2):
        codigo_prenda = str(fila[2].value or "").strip().upper()
        if codigo_prenda:
            filas_por_codigo[codigo_prenda] = fila

    hoy = datetime.now().date()
    prendas_actualizadas = []

    for item in filas_a_guardar:
        codigo_prenda = item["codigo_prenda"]
        fila = filas_por_codigo.get(codigo_prenda)

        if fila is None:
            return False, f"ERROR: NO SE ENCONTRO LA PRENDA {codigo_prenda}."

        codigo_proveedora_fila = str(fila[1].value or "").strip().upper()
        estado = str(fila[8].value or "").strip().upper()

        if codigo_proveedora_fila != codigo_proveedora:
            return False, f"ERROR: LA PRENDA {codigo_prenda} NO PERTENECE A LA PROVEEDORA {codigo_proveedora}."

        if estado != "DISPONIBLE":
            return False, f"ERROR: LA PRENDA {codigo_prenda} NO ESTA DISPONIBLE PARA REMARQUE."

        try:
            fecha_ingreso = convertir_fecha(fila[0].value)
        except ValueError:
            return False, f"ERROR: LA FECHA_INGRESO DE {codigo_prenda} ES INVALIDA."

        dias_en_inventario = (hoy - fecha_ingreso.date()).days
        if dias_en_inventario < 90:
            return False, f"ERROR: LA PRENDA {codigo_prenda} TODAVIA NO CUMPLE 3 MESES EN INVENTARIO."

        fila[encabezados["PRECIO_REMARCADO"] - 1].value = item["precio_remarcado"]
        fila[encabezados["FECHA_REMARQUE"] - 1].value = fecha_remarque
        fila[encabezados["DECISION_PROVEEDORA"] - 1].value = "PENDIENTE"

        prendas_actualizadas.append({
            "codigo_prenda": codigo_prenda,
            "articulo": str(fila[3].value or "").strip(),
            "precio_actual": fila[7].value if isinstance(fila[7].value, (int, float)) else str(fila[7].value or "").strip(),
            "precio_remarcado": item["precio_remarcado"],
            "dias_en_inventario": dias_en_inventario,
            "decision_proveedora": "PENDIENTE",
        })

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, {
        "codigo_proveedora": codigo_proveedora,
        "cantidad_actualizadas": len(prendas_actualizadas),
        "fecha_remarque": fecha_remarque,
        "prendas_actualizadas": prendas_actualizadas,
    }


def exportar_lote_remarque_excel(codigo_proveedora, nombre_proveedora, prendas_remarque, fecha_remarque=None):
    codigo_proveedora = limpiar_texto(str(codigo_proveedora))
    nombre_proveedora = str(nombre_proveedora or "").strip()

    if not codigo_proveedora:
        return False, "ERROR: CODIGO_PROVEEDORA es obligatorio."

    if fecha_remarque is None:
        fecha_remarque = datetime.now().strftime("%d/%m/%Y")
    else:
        fecha_remarque = str(fecha_remarque).strip()

    try:
        datetime.strptime(fecha_remarque, "%d/%m/%Y")
    except ValueError:
        return False, "ERROR: FECHA_REMARQUE invalida. Usa el formato DD/MM/YYYY."

    filas_exportacion = []
    for fila in prendas_remarque:
        codigo_prenda = limpiar_texto(str(fila.get("codigo_prenda", "")))
        precio_texto = str(fila.get("precio_remarcado", "")).strip()

        if not codigo_prenda and not precio_texto:
            continue

        if not codigo_prenda or not precio_texto:
            continue

        try:
            precio_remarcado = int(precio_texto.replace(".", "").replace(",", ""))
        except ValueError:
            return False, f"ERROR: PRECIO_REMARCADO invalido para {codigo_prenda}."

        filas_exportacion.append({
            "codigo_prenda": codigo_prenda,
            "articulo": str(fila.get("articulo", "")).strip(),
            "marca": str(fila.get("marca", "")).strip(),
            "talle": str(fila.get("talle", "")).strip(),
            "color": str(fila.get("color", "")).strip(),
            "precio_actual": fila.get("precio_actual", ""),
            "precio_remarcado": precio_remarcado,
        })

    if not filas_exportacion:
        return False, "ERROR: No hay prendas remarcadas para exportar."

    try:
        wb_export = Workbook()
        ws_export = wb_export.active
        ws_export.title = "REMARQUE"

        ws_export.append(["LOTE DE REMARQUE"])
        ws_export.append(["FECHA REMARQUE", fecha_remarque])
        ws_export.append(["CODIGO PROVEEDORA", codigo_proveedora])
        ws_export.append(["NOMBRE PROVEEDORA", nombre_proveedora])
        ws_export.append(["CANTIDAD PRENDAS", len(filas_exportacion)])
        ws_export.append([])
        ws_export.append([
            "CODIGO PRENDA",
            "ARTICULO",
            "MARCA",
            "TALLE",
            "COLOR",
            "PRECIO ACTUAL",
            "PRECIO REMARCADO",
            "DECISION PROVEEDORA",
        ])

        for fila in filas_exportacion:
            ws_export.append([
                fila["codigo_prenda"],
                fila["articulo"],
                fila["marca"],
                fila["talle"],
                fila["color"],
                fila["precio_actual"],
                fila["precio_remarcado"],
                "",
            ])

        formato_miles = "#,##0"
        for fila in ws_export.iter_rows(min_row=8, max_row=7 + len(filas_exportacion), min_col=6, max_col=7):
            for celda in fila:
                if isinstance(celda.value, (int, float)):
                    celda.number_format = formato_miles

        for columna in ws_export.columns:
            largo_maximo = 0
            letra_columna = get_column_letter(columna[0].column)

            for celda in columna:
                valor = "" if celda.value is None else str(celda.value)
                if len(valor) > largo_maximo:
                    largo_maximo = len(valor)

            ws_export.column_dimensions[letra_columna].width = largo_maximo + 2

        fecha_obj = datetime.strptime(fecha_remarque, "%d/%m/%Y")
        carpeta_base = obtener_ruta_en_base("Remarques Enviados")
        carpeta_mes = os.path.join(carpeta_base, fecha_obj.strftime("%Y-%m"))
        os.makedirs(carpeta_mes, exist_ok=True)

        nombre_archivo = f"remarque_{codigo_proveedora}_{fecha_obj.strftime('%d-%m-%Y')}.xlsx"
        ruta_archivo = os.path.join(carpeta_mes, nombre_archivo)
        wb_export.save(ruta_archivo)

        return True, ruta_archivo
    except PermissionError:
        return False, "ERROR: No se pudo guardar el archivo Excel de remarque porque esta abierto."


def obtener_lote_decision_remarque_desde_gui(codigo_proveedora):
    codigo_proveedora = limpiar_texto(str(codigo_proveedora))

    if not codigo_proveedora:
        return False, "ERROR: CODIGO_PROVEEDORA es obligatorio."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    nombre_proveedora = ""
    if "PROVEEDORAS" in wb.sheetnames:
        ws_proveedoras = wb["PROVEEDORAS"]
        for fila in ws_proveedoras.iter_rows(min_row=2, values_only=True):
            if str(fila[0] or "").strip().upper() == codigo_proveedora:
                nombre_proveedora = str(fila[1] or "").strip()
                break

    encabezados = _asegurar_columnas_remarque(ws_ingresos)
    prendas_pendientes = []

    for fila in ws_ingresos.iter_rows(min_row=2):
        codigo_proveedora_fila = str(fila[1].value or "").strip().upper()
        estado = str(fila[8].value or "").strip().upper()
        decision = str(fila[encabezados["DECISION_PROVEEDORA"] - 1].value or "").strip().upper()
        precio_remarcado = fila[encabezados["PRECIO_REMARCADO"] - 1].value

        if codigo_proveedora_fila != codigo_proveedora:
            continue
        if estado != "DISPONIBLE":
            continue
        if decision != "PENDIENTE":
            continue
        if precio_remarcado in ("", None):
            continue

        prendas_pendientes.append({
            "codigo_prenda": str(fila[2].value or "").strip(),
            "articulo": str(fila[3].value or "").strip(),
            "marca": str(fila[4].value or "").strip(),
            "talle": str(fila[5].value or "").strip(),
            "color": str(fila[6].value or "").strip(),
            "precio_actual": fila[7].value if isinstance(fila[7].value, (int, float)) else str(fila[7].value or "").strip(),
            "precio_remarcado": precio_remarcado,
            "fecha_remarque": fila[encabezados["FECHA_REMARQUE"] - 1].value,
            "decision_proveedora": decision,
        })

    prendas_pendientes.sort(key=lambda fila: fila["codigo_prenda"])

    if not prendas_pendientes:
        return False, "NO HAY REMARQUES PENDIENTES PARA ESTA PROVEEDORA."

    return True, {
        "codigo_proveedora": codigo_proveedora,
        "nombre_proveedora": nombre_proveedora,
        "cantidad_prendas": len(prendas_pendientes),
        "prendas": prendas_pendientes,
    }


def guardar_decisiones_remarque_desde_gui(codigo_proveedora, decisiones):
    codigo_proveedora = limpiar_texto(str(codigo_proveedora))

    if not codigo_proveedora:
        return False, "ERROR: CODIGO_PROVEEDORA es obligatorio."

    decisiones_validas = []
    codigos_vistos = set()

    for fila in decisiones:
        codigo_prenda = limpiar_texto(str(fila.get("codigo_prenda", "")))
        decision = limpiar_texto(str(fila.get("decision", "")))

        if not codigo_prenda and not decision:
            continue
        if not codigo_prenda:
            return False, "ERROR: HAY UNA FILA SIN CODIGO_PRENDA."
        if decision not in {"APROBADO", "DEVOLVER", "PENDIENTE"}:
            return False, f"ERROR: DECISION invalida para {codigo_prenda}."
        if codigo_prenda in codigos_vistos:
            return False, f"ERROR: CODIGO_PRENDA duplicado en decisiones: {codigo_prenda}."

        decisiones_validas.append({
            "codigo_prenda": codigo_prenda,
            "decision": decision,
        })
        codigos_vistos.add(codigo_prenda)

    if not decisiones_validas:
        return False, "ERROR: No hay decisiones para guardar."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    encabezados = _asegurar_columnas_remarque(ws_ingresos)
    filas_por_codigo = {}
    for fila in ws_ingresos.iter_rows(min_row=2):
        codigo = str(fila[2].value or "").strip().upper()
        if codigo:
            filas_por_codigo[codigo] = fila

    prendas_aprobadas = []
    prendas_devolucion = []

    for item in decisiones_validas:
        codigo_prenda = item["codigo_prenda"]
        decision = item["decision"]
        fila = filas_por_codigo.get(codigo_prenda)

        if fila is None:
            return False, f"ERROR: NO SE ENCONTRO LA PRENDA {codigo_prenda}."

        codigo_proveedora_fila = str(fila[1].value or "").strip().upper()
        estado = str(fila[8].value or "").strip().upper()
        decision_actual = str(fila[encabezados["DECISION_PROVEEDORA"] - 1].value or "").strip().upper()
        precio_remarcado = fila[encabezados["PRECIO_REMARCADO"] - 1].value

        if codigo_proveedora_fila != codigo_proveedora:
            return False, f"ERROR: LA PRENDA {codigo_prenda} NO PERTENECE A LA PROVEEDORA {codigo_proveedora}."
        if estado != "DISPONIBLE":
            return False, f"ERROR: LA PRENDA {codigo_prenda} NO ESTA DISPONIBLE PARA DECIDIR."
        if decision_actual != "PENDIENTE":
            return False, f"ERROR: LA PRENDA {codigo_prenda} YA NO TIENE DECISION PENDIENTE."
        if precio_remarcado in ("", None):
            return False, f"ERROR: LA PRENDA {codigo_prenda} NO TIENE PRECIO_REMARCADO."

        if decision == "PENDIENTE":
            continue

        if decision == "APROBADO":
            precio_anterior = fila[7].value
            fila[7].value = precio_remarcado
            fila[encabezados["DECISION_PROVEEDORA"] - 1].value = "APROBADO"
            prendas_aprobadas.append({
                "codigo_prenda": codigo_prenda,
                "precio_anterior": precio_anterior,
                "precio_nuevo": precio_remarcado,
            })
            continue

        fila[8].value = "PENDIENTE DEVOLUCION"
        fila[encabezados["DECISION_PROVEEDORA"] - 1].value = "NO APROBADO"
        prendas_devolucion.append({
            "codigo_prenda": codigo_prenda,
            "precio_remarcado": precio_remarcado,
        })

    if not prendas_aprobadas and not prendas_devolucion:
        return False, "ERROR: No hay decisiones nuevas para aplicar."

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, {
        "codigo_proveedora": codigo_proveedora,
        "cantidad_aprobadas": len(prendas_aprobadas),
        "cantidad_devolucion": len(prendas_devolucion),
        "prendas_aprobadas": prendas_aprobadas,
        "prendas_devolucion": prendas_devolucion,
    }


def validar_venta_pendiente_por_codigo(codigo_prenda):
    codigo_prenda = limpiar_texto(str(codigo_prenda))

    if not codigo_prenda:
        return False, "ERROR: CODIGO_PRENDA es obligatorio."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS o VENTAS en el archivo Excel."

    venta_encontrada = False

    for fila in ws_ventas.iter_rows(min_row=2):
        if str(fila[2].value or "").strip().upper() == codigo_prenda:
            venta_encontrada = True
            validacion_actual = str(fila[11].value or "").strip().upper()

            if validacion_actual == "PAGADO":
                return False, "ERROR: ESTA VENTA YA ESTA VALIDADA COMO PAGADO."

            fila[11].value = "PAGADO"
            break

    if not venta_encontrada:
        return False, "ERROR: NO SE ENCONTRO LA VENTA EN VENTAS."

    for fila in ws_ingresos.iter_rows(min_row=2):
        if str(fila[2].value or "").strip().upper() == codigo_prenda:
            fila[13].value = "PAGADO"
            break

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, f"VENTA {codigo_prenda} VALIDADA CORRECTAMENTE."


                                        # VENTAS PENDIENTES DE VALIDACION
def ventas_pendientes_validacion():
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return
    except FileNotFoundError:
        print("ERROR: No se encontro el archivo fashion_reset.xlsx.")
        return
    except KeyError:
        print("ERROR: No existe la hoja VENTAS en el archivo Excel.")
        return

    print("\n=== VENTAS PENDIENTES DE VALIDACION ===")

    cantidad_pendientes = 0
    total_pendiente = 0

    for fila in ws_ventas.iter_rows(min_row=2, values_only=True):
        fecha_venta = fila[0]
        codigo_proveedora = fila[1]
        codigo_prenda = fila[2]
        articulo = fila[3]
        precio_venta = fila[8]
        cliente = fila[9]
        tipo_pago = fila[10]
        validacion = str(fila[11]).strip().upper()

        if validacion == "PENDIENTE":
            cantidad_pendientes += 1

            if isinstance(precio_venta, int):
                total_pendiente += precio_venta

            print(f"""
FECHA_VENTA: {fecha_venta}
CODIGO_PROVEEDORA: {codigo_proveedora}
CODIGO_PRENDA: {codigo_prenda}
ARTICULO: {articulo}
PRECIO_VENTA: {precio_venta}
CLIENTE: {cliente}
TIPO_PAGO: {tipo_pago}
VALIDACION: {validacion}
""")

    if cantidad_pendientes == 0:
        print("NO HAY VENTAS PENDIENTES DE VALIDACION.")
        return

    print(f"""
=== RESUMEN ===
CANTIDAD_PENDIENTES: {cantidad_pendientes}
TOTAL_IMPORTE_PENDIENTE: {total_pendiente}
""")



                                            # VALIDAR VENTA PENDIENTE
def validar_venta_pendiente():
    print("\n=== VALIDAR VENTA PENDIENTE ===")

    codigo_prenda = limpiar_texto(input("CODIGO_PRENDA: "))

    if not codigo_prenda:
        print("ERROR: CODIGO_PRENDA es obligatorio.")
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return
    except FileNotFoundError:
        print("ERROR: No se encontro el archivo fashion_reset.xlsx.")
        return
    except KeyError:
        print("ERROR: No existe la hoja INGRESOS o VENTAS en el archivo Excel.")
        return

    venta_encontrada = False

    for fila in ws_ventas.iter_rows(min_row=2):
        if str(fila[2].value).strip().upper() == codigo_prenda:
            venta_encontrada = True

            validacion_actual = str(fila[11].value).strip().upper()

            if validacion_actual == "PAGADO":
                print("ERROR: ESTA VENTA YA ESTA VALIDADA COMO PAGADO.")
                return

            print("\n=== DATOS DE LA VENTA ===")
            print(f"FECHA_VENTA: {fila[0].value}")
            print(f"CODIGO_PROVEEDORA: {fila[1].value}")
            print(f"CODIGO_PRENDA: {fila[2].value}")
            print(f"ARTICULO: {fila[3].value}")
            print(f"PRECIO_VENTA: {fila[8].value}")
            print(f"CLIENTE: {fila[9].value}")
            print(f"TIPO_PAGO: {fila[10].value}")
            print(f"VALIDACION ACTUAL: {fila[11].value}")

            confirmar = input("CONFIRMAR CAMBIAR A PAGADO? (S/N): ").strip().upper()
            if confirmar != "S":
                print("OPERACION CANCELADA.")
                return

            fila[11].value = "PAGADO"
            break

    if venta_encontrada == False:
        print("ERROR: NO SE ENCONTRO LA VENTA EN VENTAS.")
        return

    for fila in ws_ingresos.iter_rows(min_row=2):
        if str(fila[2].value).strip().upper() == codigo_prenda:
            fila[13].value = "PAGADO"
            break

    wb.save(ARCHIVO_EXCEL)
    print("VENTA VALIDADA CORRECTAMENTE.")



def calcular_rendicion_proveedora(mes_texto, anio_texto, codigo_proveedora):
    codigo_proveedora = limpiar_texto(str(codigo_proveedora))

    try:
        mes = int(str(mes_texto).strip())
        if mes < 1 or mes > 12:
            return False, "ERROR: El MES debe estar entre 1 y 12."
    except ValueError:
        return False, "ERROR: MES invalido."

    try:
        anio = int(str(anio_texto).strip())
    except ValueError:
        return False, "ERROR: AÑO invalido."

    if not codigo_proveedora:
        return False, "ERROR: CODIGO_PROVEEDORA es obligatorio."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja VENTAS en el archivo Excel."

    nombre_proveedora = ""
    if "PROVEEDORAS" in wb.sheetnames:
        ws_proveedoras = wb["PROVEEDORAS"]
        for fila in ws_proveedoras.iter_rows(min_row=2, values_only=True):
            if str(fila[0] or "").strip().upper() == codigo_proveedora:
                nombre_proveedora = str(fila[1] or "").strip().upper()
                break

    total_vendido = 0
    total_descuentos = 0
    cantidad_prendas = 0
    ventas_rendicion = []

    for fila in ws_ventas.iter_rows(min_row=2, values_only=True):
        fecha_venta = fila[0]
        codigo_proveedora_fila = str(fila[1] or "").strip().upper()
        codigo_prenda = fila[2]
        articulo = fila[3]
        precio_venta = fila[8]
        cliente = fila[9]
        tipo_pago = fila[10]
        validacion = fila[11]

        try:
            fecha_obj = convertir_fecha(fecha_venta)
        except Exception:
            continue

        if (
            fecha_obj.month == mes
            and fecha_obj.year == anio
            and codigo_proveedora_fila == codigo_proveedora
            and str(validacion).strip().upper() == "PAGADO"
        ):
            if not isinstance(precio_venta, (int, float)):
                continue

            cantidad_prendas += 1
            total_vendido += int(precio_venta)
            ventas_rendicion.append({
                "fecha_venta": fecha_venta,
                "codigo_proveedora": codigo_proveedora_fila,
                "codigo_prenda": codigo_prenda,
                "articulo": articulo,
                "precio_venta": int(precio_venta),
                "cliente": cliente,
                "tipo_pago": tipo_pago,
                "validacion": validacion
            })

    for fila in ws_ventas.iter_rows(min_row=2, values_only=True):
        fecha_venta = fila[0]
        precio_venta = fila[8]
        tipo_pago = fila[10]
        validacion = fila[11]
        obs_venta = str(fila[12] or "").strip().upper()

        try:
            fecha_obj = convertir_fecha(fecha_venta)
        except Exception:
            continue

        if fecha_obj.month == mes and fecha_obj.year == anio and str(validacion).strip().upper() == "PAGADO":
            if str(tipo_pago).strip().upper() == "DESCUENTO A PROVEEDORA":
                texto_esperado = f"DESCUENTO_PROVEEDORA: {codigo_proveedora}"
                if obs_venta == texto_esperado and isinstance(precio_venta, (int, float)):
                    total_descuentos += int(precio_venta)

    if cantidad_prendas == 0:
        return False, "NO SE ENCONTRARON VENTAS PAGADAS PARA ESTA PROVEEDORA EN ESE PERIODO."

    comision_proveedora = int(total_vendido * 0.60)
    comision_fashion_reset = int(total_vendido * 0.40)
    saldo_final = int(comision_proveedora - total_descuentos)

    return True, {
        "mes": mes,
        "anio": anio,
        "codigo_proveedora": codigo_proveedora,
        "nombre_proveedora": nombre_proveedora,
        "cantidad_prendas": cantidad_prendas,
        "total_vendido": total_vendido,
        "comision_proveedora": comision_proveedora,
        "comision_fashion_reset": comision_fashion_reset,
        "total_descuentos": total_descuentos,
        "saldo_final": saldo_final,
        "ventas_rendicion": ventas_rendicion
    }


                                                # RENDICION DE CUENTAS
def rendicion_cuentas():
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return

    print("\n=== RENDICION DE CUENTAS ===")

    # PEDIR DATOS DE BUSQUEDA
    mes_texto = input("MES (1-12): ").strip()
    anio_texto = input("AÑO: ").strip()
    codigo_proveedora = limpiar_texto(input("CODIGO_PROVEEDORA: "))

    # VALIDAR MES
    try:
        mes = int(mes_texto)
        if mes < 1 or mes > 12:
            print("ERROR: El MES debe estar entre 1 y 12.")
            return
    except ValueError:
        print("ERROR: MES invalido.")
        return

    # VALIDAR AÑO
    try:
        anio = int(anio_texto)
    except ValueError:
        print("ERROR: AÑO invalido.")
        return

    if not codigo_proveedora:
        print("ERROR: CODIGO_PROVEEDORA es obligatorio.")
        return

    print(f"""
BUSCANDO RENDICION...
MES: {mes}
AÑO: {anio}
CODIGO_PROVEEDORA: {codigo_proveedora}
""")
    # VARIABLES DE RENDICION
    total_vendido = 0
    total_descuentos = 0
    cantidad_prendas = 0
    ventas_rendicion = []

    print("\n=== DETALLE DE VENTAS ===")

    # RECORRER HOJA VENTAS
    for fila in ws_ventas.iter_rows(min_row=2, values_only=True):
        fecha_venta = fila[0]
        codigo_proveedora_fila = str(fila[1]).strip().upper()
        codigo_prenda = fila[2]
        articulo = fila[3]
        precio_venta = fila[8]
        cliente = fila[9]
        tipo_pago = fila[10]
        validacion = fila[11]
        obs_venta = str(fila[12]).strip().upper()

        # VALIDAR FECHA
        try:
            fecha_obj = convertir_fecha(fecha_venta)
        except:
            continue

        # FILTRAR MES, AÑO, PROVEEDORA Y PAGADO
        if (
            fecha_obj.month == mes
            and fecha_obj.year == anio
            and codigo_proveedora_fila == codigo_proveedora
            and str(validacion).strip().upper() == "PAGADO"
        ):
            cantidad_prendas += 1
            total_vendido += precio_venta
            
            ventas_rendicion.append([
                fecha_venta,
                codigo_proveedora_fila,
                codigo_prenda,
                articulo,
                precio_venta,
                cliente,
                tipo_pago,
                validacion
            ])


            print(f"""
CODIGO_PRENDA: {codigo_prenda}
ARTICULO: {articulo}
PRECIO_VENTA: {precio_venta}
CLIENTE: {cliente}
TIPO_PAGO: {tipo_pago}
VALIDACION: {validacion}
""")    
    
    # CALCULAR DESCUENTOS A LA PROVEEDORA DESDE TODAS LAS VENTAS DEL MES
    for fila in ws_ventas.iter_rows(min_row=2, values_only=True):
        fecha_venta = fila[0]
        precio_venta = fila[8]
        tipo_pago = fila[10]
        validacion = fila[11]
        obs_venta = str(fila[12]).strip().upper()

        try:
            fecha_obj = convertir_fecha(fecha_venta)
        except:
            continue

        if fecha_obj.month == mes and fecha_obj.year == anio and str(validacion).strip().upper() == "PAGADO":
            if str(tipo_pago).strip().upper() == "DESCUENTO A PROVEEDORA":
                texto_esperado = f"DESCUENTO_PROVEEDORA: {codigo_proveedora}"
                if obs_venta == texto_esperado:
                    total_descuentos += precio_venta


    # VALIDAR SI NO HAY VENTAS
    if cantidad_prendas == 0:
        print("NO SE ENCONTRARON VENTAS PAGADAS PARA ESTA PROVEEDORA EN ESE PERIODO.")
        return

    # CALCULAR COMISIONES
    comision_proveedora = int(total_vendido * 0.60)
    comision_fashion_reset = int(total_vendido * 0.40)
    saldo_final = int(comision_proveedora - total_descuentos)

    # MOSTRAR RESUMEN FINAL
    print(f"""
=== RESUMEN DE RENDICION ===
CANTIDAD_PRENDAS: {cantidad_prendas}
TOTAL_VENDIDO: {total_vendido}
COMISION_PROVEEDORA (60%): {comision_proveedora}
COMISION_FASHION_RESET (40%): {comision_fashion_reset}
TOTAL_DESCUENTOS_A_PROVEEDORA: {total_descuentos}
SALDO_FINAL_A_PAGAR: {saldo_final}
""")

    exportar = input("¿EXPORTAR ESTA RENDICION A EXCEL? (S/N): ").strip().upper()

    if exportar == "S":
        exportar_rendicion_excel(
            mes,
            anio,
            codigo_proveedora,
            ventas_rendicion,
            cantidad_prendas,
            total_vendido,
            comision_proveedora,
            comision_fashion_reset,
            total_descuentos,
            saldo_final
        )


# EXPORTAR RENDICION A EXCEL
def exportar_rendicion_excel(mes, anio, codigo_proveedora, ventas_rendicion, cantidad_prendas, total_vendido, comision_proveedora, comision_fashion_reset, total_descuentos, saldo_final):
    try:
        wb_export = Workbook()
        ws_export = wb_export.active
        ws_export.title = "RENDICION"

        nombre_proveedora = ""
        descuentos_proveedora = []

        wb_datos = load_workbook(ARCHIVO_EXCEL)
        ws_proveedoras = wb_datos["PROVEEDORAS"]
        ws_ventas = wb_datos["VENTAS"]

        for fila in ws_proveedoras.iter_rows(min_row=2, values_only=True):
            if str(fila[0]).strip().upper() == codigo_proveedora:
                nombre_proveedora = str(fila[1]).strip().upper()
                break   

        for fila in ws_ventas.iter_rows(min_row=2, values_only=True):
            fecha_venta = fila[0]
            codigo_prenda = fila[2]
            articulo = fila[3]
            precio_venta = fila[8]
            cliente = fila[9]
            tipo_pago = str(fila[10] or "").strip().upper()
            validacion = str(fila[11] or "").strip().upper()
            obs_venta = str(fila[12] or "").strip().upper()

            try:
                fecha_obj = convertir_fecha(fecha_venta)
            except Exception:
                continue

            if (
                fecha_obj.month == int(mes)
                and fecha_obj.year == int(anio)
                and validacion == "PAGADO"
                and tipo_pago == "DESCUENTO A PROVEEDORA"
                and obs_venta == f"DESCUENTO_PROVEEDORA: {str(codigo_proveedora).strip().upper()}"
            ):
                descuentos_proveedora.append([
                    codigo_prenda,
                    articulo,
                    precio_venta
                ])

        ws_export.append(["RENDICION DE CUENTAS"])
        ws_export.append(["MES", mes])
        ws_export.append(["AÑO", anio])
        ws_export.append(["CODIGO PROVEEDORA", codigo_proveedora])
        ws_export.append(["NOMBRE PROVEEDORA", nombre_proveedora])
        ws_export.append([])

        ws_export.append([
            "CODIGO PRENDA",
            "ARTICULO",
            "PRECIO VENTA"
        ])

        for venta in ventas_rendicion:
            ws_export.append(venta)

        ws_export.append([])
        ws_export.append(["CANTIDAD PRENDAS", cantidad_prendas])
        ws_export.append(["TOTAL VENDIDO", total_vendido])
        ws_export.append(["COMISION PROVEEDORA (60%)", comision_proveedora])
        ws_export.append(["COMISION FASHION RESET (40%)", comision_fashion_reset])
        ws_export.append(["TOTAL DESCUENTOS A PROVEEDORA", total_descuentos])
        ws_export.append(["SALDO FINAL A PAGAR", saldo_final])

        if descuentos_proveedora:
            ws_export.append([])
            ws_export.append(["DETALLE DE DESCUENTOS A PROVEEDORA"])
            ws_export.append([
                "CODIGO PRENDA",
                "ARTICULO",
                "PRECIO DESCONTADO"
            ])

            for descuento in descuentos_proveedora:
                ws_export.append(descuento)

        formato_miles = "#,##0"

        for fila in ws_export.iter_rows(min_row=8, max_col=3):
            celda_precio = fila[2]
            if isinstance(celda_precio.value, (int, float)):
                celda_precio.number_format = formato_miles

        for fila in ws_export.iter_rows(min_row=1, max_col=2):
            etiqueta = str(fila[0].value or "").strip().upper()
            if etiqueta in {
                "TOTAL VENDIDO",
                "COMISION PROVEEDORA (60%)",
                "COMISION FASHION RESET (40%)",
                "TOTAL DESCUENTOS A PROVEEDORA",
                "SALDO FINAL A PAGAR",
            } and isinstance(fila[1].value, (int, float)):
                fila[1].number_format = formato_miles

        for fila in ws_export.iter_rows(min_row=1, max_col=3):
            etiqueta = str(fila[0].value or "").strip().upper()
            if etiqueta == "CODIGO PRENDA":
                continue
            if str(fila[0].value or "").strip().upper() == "" and str(fila[1].value or "").strip().upper() == "DETALLE DE DESCUENTOS A PROVEEDORA":
                continue
            if len(fila) >= 3 and isinstance(fila[2].value, (int, float)):
                fila[2].number_format = formato_miles

        for columna in ws_export.columns:
            largo_maximo = 0
            letra_columna = get_column_letter(columna[0].column)

            for celda in columna:
                valor = "" if celda.value is None else str(celda.value)
                if len(valor) > largo_maximo:
                    largo_maximo = len(valor)

            ws_export.column_dimensions[letra_columna].width = largo_maximo + 2

        carpeta_base = obtener_ruta_en_base("Rendiciones de Cuenta")
        carpeta_mes = os.path.join(carpeta_base, f"{anio}-{int(mes):02d}")
        os.makedirs(carpeta_mes, exist_ok=True)

        nombre_archivo = f"rendicion_{codigo_proveedora}_{mes}_{anio}.xlsx"
        ruta_archivo = os.path.join(carpeta_mes, nombre_archivo)
        wb_export.save(ruta_archivo)

        print(f"ARCHIVO EXCEL GENERADO: {ruta_archivo}")
        return True, ruta_archivo

    except PermissionError:
        print("ERROR: No se pudo guardar el archivo Excel de rendicion porque esta abierto.")
        return False, "ERROR: No se pudo guardar el archivo Excel de rendicion porque esta abierto."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja PROVEEDORAS en el archivo Excel."
                                             
                                             
                                             
                                             
                                             
def obtener_venta_desde_gui(codigo_prenda):
    codigo_prenda = limpiar_texto(str(codigo_prenda))

    if not codigo_prenda:
        return False, "ERROR: CODIGO_PRENDA es obligatorio."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja VENTAS en el archivo Excel."

    for fila in ws_ventas.iter_rows(min_row=2, values_only=True):
        if str(fila[2] or "").strip().upper() == codigo_prenda:
            return True, {
                "fecha_venta": str(fila[0] or "").strip(),
                "codigo_proveedora": str(fila[1] or "").strip(),
                "codigo_prenda": str(fila[2] or "").strip(),
                "articulo": str(fila[3] or "").strip(),
                "marca": str(fila[4] or "").strip(),
                "talle": str(fila[5] or "").strip(),
                "color": str(fila[6] or "").strip(),
                "precio_lista": fila[7] if isinstance(fila[7], (int, float)) else str(fila[7] or "").strip(),
                "precio_venta": fila[8] if isinstance(fila[8], (int, float)) else str(fila[8] or "").strip(),
                "cliente": str(fila[9] or "").strip(),
                "tipo_pago": str(fila[10] or "").strip(),
                "validacion": str(fila[11] or "").strip(),
                "obs_venta": str(fila[12] or "").strip(),
            }

    return False, "ERROR: NO SE ENCONTRO LA VENTA."


def actualizar_venta_desde_gui(
    codigo_prenda,
    fecha_venta,
    cliente,
    tipo_pago,
    validacion,
    precio_venta_texto,
    obs_venta
):
    codigo_prenda = limpiar_texto(str(codigo_prenda))
    fecha_venta = str(fecha_venta).strip()
    cliente = limpiar_texto(str(cliente))
    tipo_pago = limpiar_texto(str(tipo_pago))
    validacion = limpiar_texto(str(validacion))
    precio_venta_texto = str(precio_venta_texto).strip()
    obs_venta = normalizar_obs_descuento_proveedora(tipo_pago, obs_venta)

    if not codigo_prenda:
        return False, "ERROR: CODIGO_PRENDA es obligatorio."
    if not fecha_venta or not cliente or not tipo_pago or not validacion:
        return False, "ERROR: Fecha, cliente, tipo pago y validacion son obligatorios."
    if tipo_pago == "DESCUENTO A PROVEEDORA" and not obs_venta:
        return False, "ERROR: Debes indicar el codigo de la proveedora para el descuento."

    try:
        datetime.strptime(fecha_venta, "%d/%m/%Y")
    except ValueError:
        return False, "ERROR: FECHA_VENTA invalida. Usa el formato DD/MM/YYYY."

    try:
        precio_venta = int(precio_venta_texto.replace(".", "").replace(",", ""))
    except ValueError:
        return False, "ERROR: PRECIO_VENTA invalido. Debe ser un numero entero."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS o VENTAS en el archivo Excel."

    fila_venta = None
    for fila in ws_ventas.iter_rows(min_row=2):
        if str(fila[2].value or "").strip().upper() == codigo_prenda:
            fila_venta = fila
            break

    if fila_venta is None:
        return False, "ERROR: NO SE ENCONTRO LA VENTA."

    fila_ingreso = None
    for fila in ws_ingresos.iter_rows(min_row=2):
        if str(fila[2].value or "").strip().upper() == codigo_prenda:
            fila_ingreso = fila
            break

    if fila_ingreso is None:
        return False, "ERROR: NO SE ENCONTRO LA PRENDA EN INGRESOS."

    fila_venta[0].value = fecha_venta
    fila_venta[8].value = precio_venta
    fila_venta[9].value = cliente
    fila_venta[10].value = tipo_pago
    fila_venta[11].value = validacion
    fila_venta[12].value = obs_venta

    fila_ingreso[9].value = fecha_venta
    fila_ingreso[10].value = precio_venta
    fila_ingreso[11].value = cliente
    fila_ingreso[12].value = tipo_pago
    fila_ingreso[13].value = validacion
    fila_ingreso[15].value = obs_venta

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, "VENTA ACTUALIZADA CORRECTAMENTE."


def obtener_prenda_desde_gui(codigo_prenda):
    codigo_prenda = limpiar_texto(str(codigo_prenda))

    if not codigo_prenda:
        return False, "ERROR: CODIGO_PRENDA es obligatorio."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    for fila in ws_ingresos.iter_rows(min_row=2, values_only=True):
        if str(fila[2] or "").strip().upper() == codigo_prenda:
            return True, {
                "codigo_prenda": str(fila[2] or "").strip(),
                "codigo_proveedora": str(fila[1] or "").strip(),
                "articulo": str(fila[3] or "").strip(),
                "marca": str(fila[4] or "").strip(),
                "talle": str(fila[5] or "").strip(),
                "color": str(fila[6] or "").strip(),
                "precio": fila[7] if isinstance(fila[7], (int, float)) else str(fila[7] or "").strip(),
                "estado": str(fila[8] or "").strip(),
                "obs_ingreso": str(fila[14] or "").strip(),
            }

    return False, "ERROR: NO SE ENCONTRO LA PRENDA."


def actualizar_prenda_desde_gui(
    codigo_prenda,
    articulo,
    marca,
    talle,
    color,
    precio_texto,
    obs_ingreso
):
    codigo_prenda = limpiar_texto(str(codigo_prenda))
    articulo = limpiar_texto(str(articulo))
    marca = limpiar_texto(str(marca))
    talle = limpiar_texto(str(talle))
    color = limpiar_texto(str(color))
    obs_ingreso = limpiar_texto(str(obs_ingreso))
    precio_texto = str(precio_texto).strip()

    if not codigo_prenda:
        return False, "ERROR: CODIGO_PRENDA es obligatorio."
    if not articulo:
        return False, "ERROR: ARTICULO es obligatorio."

    try:
        precio = int(precio_texto.replace(".", "").replace(",", ""))
    except ValueError:
        return False, "ERROR: PRECIO invalido. Debe ser un numero entero."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    fila_prenda = None
    for fila in ws_ingresos.iter_rows(min_row=2):
        if str(fila[2].value or "").strip().upper() == codigo_prenda:
            fila_prenda = fila
            break

    if fila_prenda is None:
        return False, "ERROR: NO SE ENCONTRO LA PRENDA."

    fila_prenda[3].value = articulo
    fila_prenda[4].value = marca
    fila_prenda[5].value = talle
    fila_prenda[6].value = color
    fila_prenda[7].value = precio
    fila_prenda[14].value = obs_ingreso

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, "PRENDA ACTUALIZADA CORRECTAMENTE."


def registrar_devolucion_desde_gui(codigo_prenda):
    codigo_prenda = limpiar_texto(str(codigo_prenda))

    if not codigo_prenda:
        return False, "ERROR: CODIGO_PRENDA es obligatorio."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    fila_prenda = None
    for fila in ws_ingresos.iter_rows(min_row=2):
        if str(fila[2].value or "").strip().upper() == codigo_prenda:
            fila_prenda = fila
            break

    if fila_prenda is None:
        return False, "ERROR: NO SE ENCONTRO LA PRENDA."

    estado_actual = str(fila_prenda[8].value or "").strip().upper()
    if estado_actual == "VENDIDO":
        return False, "ERROR: NO SE PUEDE DEVOLVER UNA PRENDA VENDIDA."
    if estado_actual == "DEVUELTO":
        return False, "ERROR: LA PRENDA YA ESTA DEVUELTA."

    fila_prenda[8].value = "DEVUELTO"

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, {
        "codigo_prenda": str(fila_prenda[2].value or "").strip(),
        "codigo_proveedora": str(fila_prenda[1].value or "").strip(),
        "articulo": str(fila_prenda[3].value or "").strip(),
        "marca": str(fila_prenda[4].value or "").strip(),
        "talle": str(fila_prenda[5].value or "").strip(),
        "color": str(fila_prenda[6].value or "").strip(),
        "precio": fila_prenda[7].value,
        "estado": str(fila_prenda[8].value or "").strip(),
    }


                                            # EDITAR PRENDA
def editar_prenda():
    print("\n=== EDITAR PRENDA ===")

    codigo_prenda = input("CODIGO_PRENDA: ").strip().upper()

    wb = load_workbook(ARCHIVO_EXCEL)
    ws_ingresos = wb["INGRESOS"]

    prenda_encontrada = False

    for fila in ws_ingresos.iter_rows(min_row=2):
        codigo_prenda_fila = str(fila[2].value).strip().upper()

        if codigo_prenda_fila == codigo_prenda:
            prenda_encontrada = True

            print("\n=== DATOS ACTUALES ===")
            print(f"CODIGO_PRENDA: {fila[2].value}")
            print(f"ARTICULO: {fila[3].value}")
            print(f"MARCA: {fila[4].value}")
            print(f"TALLE: {fila[5].value}")
            print(f"COLOR: {fila[6].value}")
            print(f"PRECIO: {fila[7].value}")
            print(f"OBS_INGRESO: {fila[14].value}")
            break

    if prenda_encontrada == False:
        print("ERROR: NO SE ENCONTRO LA PRENDA.")
        return
    
    print("\n=== NUEVOS DATOS ===")
    nuevo_articulo = limpiar_texto(input("ARTICULO NUEVO (ENTER = NO CAMBIAR): "))
    nueva_marca = limpiar_texto(input("MARCA NUEVA (ENTER = NO CAMBIAR): "))
    nuevo_talle = limpiar_texto(input("TALLE NUEVO (ENTER = NO CAMBIAR): "))
    nuevo_color = limpiar_texto(input("COLOR NUEVO (ENTER = NO CAMBIAR): "))
    nuevo_precio_texto = input("PRECIO NUEVO (ENTER = NO CAMBIAR): ").strip()
    nueva_obs_ingreso = limpiar_texto(input("OBS_INGRESO NUEVA (ENTER = NO CAMBIAR): "))

    # MANTENER VALOR ACTUAL SI SE DEJA VACIO
    if not nuevo_articulo:
        nuevo_articulo = str(fila[3].value).strip().upper()

    if not nueva_marca:
        nueva_marca = str(fila[4].value).strip().upper()

    if not nuevo_talle:
        nuevo_talle = str(fila[5].value).strip().upper()

    if not nuevo_color:
        nuevo_color = str(fila[6].value).strip().upper()

    if not nuevo_precio_texto:
        nuevo_precio = fila[7].value
    else:
        try:
            nuevo_precio = int(nuevo_precio_texto)
        except ValueError:
            print("ERROR: PRECIO NUEVO invalido. Debe ser un numero entero.")
            return

    if not nueva_obs_ingreso:
        nueva_obs_ingreso = str(fila[14].value).strip().upper()

    # GUARDAR CAMBIOS EN INGRESOS
    fila[3].value = nuevo_articulo
    fila[4].value = nueva_marca
    fila[5].value = nuevo_talle
    fila[6].value = nuevo_color
    fila[7].value = nuevo_precio
    fila[14].value = nueva_obs_ingreso

    try:
        wb.save(ARCHIVO_EXCEL)
        print("\n=== PRENDA ACTUALIZADA CORRECTAMENTE ===")
        print(f"CODIGO_PRENDA: {fila[2].value}")
        print(f"ARTICULO: {fila[3].value}")
        print(f"MARCA: {fila[4].value}")
        print(f"TALLE: {fila[5].value}")
        print(f"COLOR: {fila[6].value}")
        print(f"PRECIO: {fila[7].value}")
        print(f"OBS_INGRESO: {fila[14].value}")
    except PermissionError:
        print("ERROR: No se pudo guardar porque el archivo Excel esta abierto.")



                                                    # ELIMINAR INGRESO
def eliminar_ingreso():
    print("\n=== ELIMINAR INGRESO ===")

    codigo_prenda = limpiar_texto(input("CODIGO_PRENDA: "))

    if not codigo_prenda:
        print("ERROR: CODIGO_PRENDA es obligatorio.")
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return

    fila_a_eliminar = None

    for fila in ws_ingresos.iter_rows(min_row=2):
        if str(fila[2].value).strip().upper() == codigo_prenda:
            estado_actual = str(fila[8].value).strip().upper()

            if estado_actual == "VENDIDO":
                print("ERROR: NO SE PUEDE ELIMINAR UN INGRESO QUE YA FUE VENDIDO.")
                return

            print("\n=== DATOS DEL INGRESO A ELIMINAR ===")
            print(f"CODIGO_PRENDA: {fila[2].value}")
            print(f"CODIGO_PROVEEDORA: {fila[1].value}")
            print(f"ARTICULO: {fila[3].value}")
            print(f"MARCA: {fila[4].value}")
            print(f"TALLE: {fila[5].value}")
            print(f"COLOR: {fila[6].value}")
            print(f"PRECIO: {fila[7].value}")
            print(f"ESTADO: {fila[8].value}")

            confirmar = input("CONFIRMAR ELIMINAR INGRESO? (S/N): ").strip().upper()
            if confirmar != "S":
                print("OPERACION CANCELADA.")
                return

            fila_a_eliminar = fila[0].row
            break

    if fila_a_eliminar is None:
        print("ERROR: NO SE ENCONTRO LA PRENDA.")
        return

    ws_ingresos.delete_rows(fila_a_eliminar, 1)
    wb.save(ARCHIVO_EXCEL)

    print("INGRESO ELIMINADO CORRECTAMENTE.")



                                                # REVERSAR VENTA
def reversar_venta():
    print("\n=== REVERSAR VENTA ===")

    codigo_prenda = limpiar_texto(input("CODIGO_PRENDA: "))

    if not codigo_prenda:
        print("ERROR: CODIGO_PRENDA es obligatorio.")
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return

    fila_ingreso = None

    for fila in ws_ingresos.iter_rows(min_row=2):
        if str(fila[2].value).strip().upper() == codigo_prenda:
            fila_ingreso = fila

            print("\n=== DATOS DE LA VENTA A REVERSAR ===")
            print(f"CODIGO_PRENDA: {fila_ingreso[2].value}")
            print(f"CODIGO_PROVEEDORA: {fila_ingreso[1].value}")
            print(f"ARTICULO: {fila_ingreso[3].value}")
            print(f"MARCA: {fila_ingreso[4].value}")
            print(f"TALLE: {fila_ingreso[5].value}")
            print(f"COLOR: {fila_ingreso[6].value}")
            print(f"PRECIO_LISTA: {fila_ingreso[7].value}")
            print(f"ESTADO: {fila_ingreso[8].value}")
            print(f"FECHA_VENTA: {fila_ingreso[9].value}")
            print(f"PRECIO_VENTA: {fila_ingreso[10].value}")
            print(f"CLIENTE: {fila_ingreso[11].value}")
            print(f"TIPO_PAGO: {fila_ingreso[12].value}")
            print(f"VALIDACION: {fila_ingreso[13].value}")
            print(f"OBS_VENTA: {fila_ingreso[15].value}")

            confirmar = input("CONFIRMAR REVERSAR VENTA? (S/N): ").strip().upper()
            if confirmar != "S":
                print("VENTA CANCELADA.")
                return

            break

    if fila_ingreso is None:
        print("ERROR: NO SE ENCONTRO LA PRENDA EN INGRESOS.")
        return

    estado_actual = str(fila_ingreso[8].value).strip().upper()

    if estado_actual != "VENDIDO":
        print("ERROR: SOLO SE PUEDE REVERSAR UNA PRENDA CON ESTADO VENDIDO.")
        return

    fila_venta_a_eliminar = None

    for fila in ws_ventas.iter_rows(min_row=2):
        if str(fila[2].value).strip().upper() == codigo_prenda:
            fila_venta_a_eliminar = fila[0].row
            break

    if fila_venta_a_eliminar is None:
        print("ERROR: NO SE ENCONTRO LA VENTA EN LA HOJA VENTAS.")
        return

    ws_ventas.delete_rows(fila_venta_a_eliminar, 1)

    fila_ingreso[8].value = "DISPONIBLE"
    fila_ingreso[9].value = ""
    fila_ingreso[10].value = ""
    fila_ingreso[11].value = ""
    fila_ingreso[12].value = ""
    fila_ingreso[13].value = ""
    fila_ingreso[15].value = ""

    wb.save(ARCHIVO_EXCEL)

    print("\n=== VENTA REVERSADA CORRECTAMENTE ===")


def _extraer_codigos_prenda_lote(codigos_texto):
    texto = str(codigos_texto or "").replace(",", "\n").replace(";", "\n").replace("\t", "\n")
    codigos = []
    vistos = set()

    for linea in texto.splitlines():
        codigo = limpiar_texto(linea)
        if not codigo or codigo in vistos:
            continue
        codigos.append(codigo)
        vistos.add(codigo)

    return codigos


def eliminar_ingresos_desde_gui(codigos_texto):
    codigos = _extraer_codigos_prenda_lote(codigos_texto)
    if not codigos:
        return False, "ERROR: Ingresá al menos un CODIGO_PRENDA."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS en el archivo Excel."

    filas_a_eliminar = []
    eliminadas = []
    errores = []

    for codigo_prenda in codigos:
        fila_encontrada = None
        for fila in ws_ingresos.iter_rows(min_row=2):
            if str(fila[2].value or "").strip().upper() == codigo_prenda:
                fila_encontrada = fila
                break

        if fila_encontrada is None:
            errores.append(f"{codigo_prenda}: NO SE ENCONTRO LA PRENDA.")
            continue

        estado_actual = str(fila_encontrada[8].value or "").strip().upper()
        if estado_actual == "VENDIDO":
            errores.append(f"{codigo_prenda}: NO SE PUEDE ELIMINAR UN INGRESO VENDIDO.")
            continue

        filas_a_eliminar.append(fila_encontrada[0].row)
        eliminadas.append({
            "codigo_prenda": codigo_prenda,
            "articulo": str(fila_encontrada[3].value or "").strip(),
            "estado": estado_actual,
        })

    if not filas_a_eliminar and errores:
        return False, "\n".join(errores)

    for fila_numero in sorted(filas_a_eliminar, reverse=True):
        ws_ingresos.delete_rows(fila_numero, 1)

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, {
        "cantidad_eliminadas": len(eliminadas),
        "eliminadas": eliminadas,
        "errores": errores,
    }


def reversar_ventas_desde_gui(codigos_texto):
    codigos = _extraer_codigos_prenda_lote(codigos_texto)
    if not codigos:
        return False, "ERROR: Ingresá al menos un CODIGO_PRENDA."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
        ws_ventas = wb["VENTAS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja INGRESOS o VENTAS en el archivo Excel."

    ventas_a_eliminar = []
    reversadas = []
    errores = []

    for codigo_prenda in codigos:
        fila_ingreso = None
        for fila in ws_ingresos.iter_rows(min_row=2):
            if str(fila[2].value or "").strip().upper() == codigo_prenda:
                fila_ingreso = fila
                break

        if fila_ingreso is None:
            errores.append(f"{codigo_prenda}: NO SE ENCONTRO LA PRENDA EN INGRESOS.")
            continue

        estado_actual = str(fila_ingreso[8].value or "").strip().upper()
        if estado_actual != "VENDIDO":
            errores.append(f"{codigo_prenda}: SOLO SE PUEDE REVERSAR UNA PRENDA VENDIDA.")
            continue

        fila_venta = None
        for fila in ws_ventas.iter_rows(min_row=2):
            if str(fila[2].value or "").strip().upper() == codigo_prenda:
                fila_venta = fila
                break

        if fila_venta is None:
            errores.append(f"{codigo_prenda}: NO SE ENCONTRO LA VENTA EN LA HOJA VENTAS.")
            continue

        ventas_a_eliminar.append(fila_venta[0].row)

        fila_ingreso[8].value = "DISPONIBLE"
        fila_ingreso[9].value = ""
        fila_ingreso[10].value = ""
        fila_ingreso[11].value = ""
        fila_ingreso[12].value = ""
        fila_ingreso[13].value = ""
        fila_ingreso[15].value = ""

        reversadas.append({
            "codigo_prenda": codigo_prenda,
            "articulo": str(fila_ingreso[3].value or "").strip(),
        })

    if not ventas_a_eliminar and errores:
        return False, "\n".join(errores)

    for fila_numero in sorted(ventas_a_eliminar, reverse=True):
        ws_ventas.delete_rows(fila_numero, 1)

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, {
        "cantidad_reversadas": len(reversadas),
        "reversadas": reversadas,
        "errores": errores,
    }
    print(f"CODIGO_PRENDA: {fila_ingreso[2].value}")
    print(f"ESTADO NUEVO: {fila_ingreso[8].value}")




                                # PRENDAS CON 3 MESES O MAS
def detectar_prendas_3_meses():
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return
  

    print("\n=== PRENDAS CON 3 MESES O MAS ===")

    codigo_proveedora_filtro = limpiar_texto(input("CODIGO_PROVEEDORA (ENTER = TODAS): "))
    fecha_hoy = datetime.now()
    cantidad_encontradas = 0

    for fila in ws_ingresos.iter_rows(min_row=2, values_only=True):
        fecha_ingreso = fila[0]
        codigo_proveedora = fila[1]
        codigo_prenda = fila[2]
        articulo = fila[3]
        marca = fila[4]
        talle = fila[5]
        color = fila[6]
        precio = fila[7]
        estado = fila[8]

        if str(estado).strip().upper() != "DISPONIBLE":
            continue

        try:
            fecha_obj = convertir_fecha(fecha_ingreso)
        except:
            continue

        dias_transcurridos = (fecha_hoy - fecha_obj).days

        if codigo_proveedora_filtro and str(codigo_proveedora).strip().upper() != codigo_proveedora_filtro:
            continue
    
        if dias_transcurridos >= 90:
            cantidad_encontradas += 1

            print(f"""
CODIGO_PROVEEDORA: {codigo_proveedora}
CODIGO_PRENDA: {codigo_prenda}
FECHA_INGRESO: {fecha_ingreso}
DIAS_EN_STOCK: {dias_transcurridos}
ARTICULO: {articulo}
MARCA: {marca}
TALLE: {talle}
COLOR: {color}
PRECIO: {precio}
ESTADO: {estado}
""")

    if cantidad_encontradas == 0:
        if codigo_proveedora_filtro:
            print(f"NO SE ENCONTRARON PRENDAS DISPONIBLES CON 3 MESES O MAS PARA LA PROVEEDORA {codigo_proveedora_filtro}.")
        else:
            print("NO SE ENCONTRARON PRENDAS DISPONIBLES CON 3 MESES O MAS.")
        return

    print(f"TOTAL DE PRENDAS ENCONTRADAS: {cantidad_encontradas}")
                                           
                                
                                
                                # PROPONER REMARQUE DE PRECIO
def proponer_remarque():
    print("\n=== PROPONER REMARQUE ===")

    codigo_prenda = limpiar_texto(input("CODIGO_PRENDA: "))
    nuevo_precio_texto = input("NUEVO PRECIO PROPUESTO: ").strip()

    if not codigo_prenda:
        print("ERROR: CODIGO_PRENDA es obligatorio.")
        return

    try:
        nuevo_precio = int(nuevo_precio_texto)
    except ValueError:
        print("ERROR: NUEVO PRECIO invalido.")
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return

    prenda_encontrada = False

    for fila in ws_ingresos.iter_rows(min_row=2):
        if str(fila[2].value).strip().upper() == codigo_prenda:
            prenda_encontrada = True

            if str(fila[8].value).strip().upper() != "DISPONIBLE":
                print("ERROR: SOLO SE PUEDE PROPONER REMARQUE PARA PRENDAS DISPONIBLES.")
                return

            fila[14].value = f"REMARQUE PENDIENTE - NUEVO PRECIO: {nuevo_precio}"

            wb.save(ARCHIVO_EXCEL)

            print("\n=== REMARQUE REGISTRADO ===")
            print(f"CODIGO_PRENDA: {fila[2].value}")
            print(f"PRECIO ACTUAL: {fila[7].value}")
            print(f"NUEVO PRECIO PROPUESTO: {nuevo_precio}")
            print(f"OBS_INGRESO: {fila[14].value}")
            return

    if prenda_encontrada == False:
        print("ERROR: NO SE ENCONTRO LA PRENDA.")      


                                    # APROBAR REMARQUE
def aprobar_remarque():
    print("\n=== APROBAR REMARQUE ===")

    codigo_prenda = limpiar_texto(input("CODIGO_PRENDA: "))

    if not codigo_prenda:
        print("ERROR: CODIGO_PRENDA es obligatorio.")
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return

    for fila in ws_ingresos.iter_rows(min_row=2):
        if str(fila[2].value).strip().upper() == codigo_prenda:
            obs_ingreso = str(fila[14].value).strip().upper()

            if "REMARQUE PENDIENTE - NUEVO PRECIO:" not in obs_ingreso:
                print("ERROR: LA PRENDA NO TIENE UN REMARQUE PENDIENTE.")
                return

            try:
                nuevo_precio = int(obs_ingreso.split(":")[-1].strip())
            except ValueError:
                print("ERROR: NO SE PUDO LEER EL NUEVO PRECIO PROPUESTO.")
                return

            precio_anterior = fila[7].value
            fila[7].value = nuevo_precio
            fila[14].value = f"REMARCADO APROBADO - PRECIO ANTERIOR: {precio_anterior} - PRECIO NUEVO: {nuevo_precio}"

            wb.save(ARCHIVO_EXCEL)

            print("\n=== REMARQUE APROBADO ===")
            print(f"CODIGO_PRENDA: {fila[2].value}")
            print(f"PRECIO ANTERIOR: {precio_anterior}")
            print(f"PRECIO NUEVO: {nuevo_precio}")
            print(f"OBS_INGRESO: {fila[14].value}")
            return

    print("ERROR: NO SE ENCONTRO LA PRENDA.")


                                                # DEVOLUCION DE PRENDA
             
def devolver_prendas_multiples():
    print("\n=== DEVOLUCION MULTIPLE DE PRENDAS ===")

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_ingresos = obtener_hoja(wb, "INGRESOS")
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return

    cantidad_devueltas = 0

    while True:
        codigo_prenda = limpiar_texto(input("CODIGO_PRENDA: "))

        if not codigo_prenda:
            print("ERROR: CODIGO_PRENDA es obligatorio.")
            otra = input("¿DEVOLVER OTRA PRENDA? (S/N): ").strip().upper()
            if otra != "S":
                break
            continue

        prenda_encontrada = False

        for fila in ws_ingresos.iter_rows(min_row=2):
            if str(fila[2].value).strip().upper() == codigo_prenda:
                prenda_encontrada = True
                estado_actual = str(fila[8].value).strip().upper()

                if estado_actual == "VENDIDO":
                    print("ERROR: NO SE PUEDE DEVOLVER UNA PRENDA VENDIDA.")
                    break

                if estado_actual == "DEVUELTO":
                    print("ERROR: LA PRENDA YA ESTA DEVUELTA.")
                    break

                fila[8].value = "DEVUELTO"
                cantidad_devueltas += 1

                print(f"PRENDA DEVUELTA: {fila[2].value} - {fila[3].value}")
                break

        if prenda_encontrada == False:
            print("ERROR: NO SE ENCONTRO LA PRENDA.")

        otra = input("¿DEVOLVER OTRA PRENDA? (S/N): ").strip().upper()
        if otra != "S":
            break

    if cantidad_devueltas == 0:
        print("NO SE REALIZARON DEVOLUCIONES.")
        return

    wb.save(ARCHIVO_EXCEL)
    print(f"TOTAL DE PRENDAS DEVUELTAS: {cantidad_devueltas}")
 
 
def crear_proveedora_desde_gui(
    codigo_proveedora,
    nombre_proveedora,
    telefono,
    banco,
    numero_cuenta,
    titular_cuenta,
    alias,
    obs_ingreso
):
    codigo_proveedora = limpiar_texto(str(codigo_proveedora))
    nombre_proveedora = limpiar_texto(str(nombre_proveedora))
    telefono = limpiar_texto(str(telefono))
    banco = limpiar_texto(str(banco))
    numero_cuenta = limpiar_texto(str(numero_cuenta))
    titular_cuenta = limpiar_texto(str(titular_cuenta))
    alias = limpiar_texto(str(alias))
    obs_ingreso = limpiar_texto(str(obs_ingreso))

    if not codigo_proveedora or not nombre_proveedora:
        return False, "ERROR: CODIGO_PROVEEDORA y NOMBRE_PROVEEDORA son obligatorios."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_proveedoras = wb["PROVEEDORAS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja PROVEEDORAS en el archivo Excel."

    for fila in ws_proveedoras.iter_rows(min_row=2, values_only=True):
        codigo_existente = fila[0]
        if codigo_existente and str(codigo_existente).strip().upper() == codigo_proveedora:
            return False, "ERROR: YA EXISTE UNA PROVEEDORA CON ESE CODIGO."

    ws_proveedoras.append([
        codigo_proveedora,
        nombre_proveedora,
        telefono,
        banco,
        numero_cuenta,
        titular_cuenta,
        alias,
        obs_ingreso,
        "ACTIVA"
    ])

    try:
        wb.save(ARCHIVO_EXCEL)
    except PermissionError:
        return False, "ERROR: No se pudo guardar porque el archivo Excel esta abierto."

    return True, {
        "codigo_proveedora": codigo_proveedora,
        "nombre_proveedora": nombre_proveedora,
        "estado": "ACTIVA"
    }


def obtener_proveedora_desde_gui(codigo_proveedora):
    codigo_proveedora = limpiar_texto(str(codigo_proveedora))

    if not codigo_proveedora:
        return False, "ERROR: CODIGO_PROVEEDORA es obligatorio."

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_proveedoras = wb["PROVEEDORAS"]
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja PROVEEDORAS o INGRESOS en el archivo Excel."

    datos_proveedora = None
    for fila in ws_proveedoras.iter_rows(min_row=2, values_only=True):
        if str(fila[0] or "").strip().upper() == codigo_proveedora:
            datos_proveedora = {
                "codigo_proveedora": str(fila[0] or "").strip().upper(),
                "nombre_proveedora": str(fila[1] or "").strip(),
                "telefono": str(fila[2] or "").strip(),
                "banco": str(fila[3] or "").strip(),
                "numero_cuenta": str(fila[4] or "").strip(),
                "titular_cuenta": str(fila[5] or "").strip(),
                "alias": str(fila[6] or "").strip(),
                "obs_ingreso": str(fila[7] or "").strip(),
                "estado": str(fila[8] or "").strip(),
            }
            break

    if datos_proveedora is None:
        return False, "ERROR: NO SE ENCONTRO LA PROVEEDORA."

    total_prendas = 0
    prendas_disponibles = 0
    prendas_vendidas = 0
    prendas_devueltas = 0
    detalle_prendas = []

    for fila in ws_ingresos.iter_rows(min_row=2, values_only=True):
        codigo_fila = str(fila[1] or "").strip().upper()
        estado = str(fila[8] or "").strip().upper()

        if codigo_fila == codigo_proveedora:
            total_prendas += 1

            if estado == "DISPONIBLE":
                prendas_disponibles += 1
            elif estado == "VENDIDO":
                prendas_vendidas += 1
            elif estado in {"DEVUELTO", "PENDIENTE DEVOLUCION"}:
                prendas_devueltas += 1

            detalle_prendas.append({
                "codigo_prenda": str(fila[2] or "").strip(),
                "articulo": str(fila[3] or "").strip(),
                "marca": str(fila[4] or "").strip(),
                "talle": str(fila[5] or "").strip(),
                "color": str(fila[6] or "").strip(),
                "precio": fila[7] if isinstance(fila[7], (int, float)) else str(fila[7] or "").strip(),
                "estado": str(fila[8] or "").strip(),
            })

    return True, {
        "proveedora": datos_proveedora,
        "resumen_prendas": {
            "total_prendas": total_prendas,
            "prendas_disponibles": prendas_disponibles,
            "prendas_vendidas": prendas_vendidas,
            "prendas_devueltas": prendas_devueltas,
        },
        "detalle_prendas": detalle_prendas
    }


def obtener_todas_las_proveedoras_desde_gui():
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_proveedoras = wb["PROVEEDORAS"]
    except PermissionError:
        return False, "ERROR: Cierra el archivo Excel antes de continuar."
    except FileNotFoundError:
        return False, "ERROR: No se encontro el archivo fashion_reset.xlsx."
    except KeyError:
        return False, "ERROR: No existe la hoja PROVEEDORAS en el archivo Excel."

    proveedoras = []
    for fila in ws_proveedoras.iter_rows(min_row=2, values_only=True):
        if not any(fila):
            continue

        proveedoras.append({
            "codigo_proveedora": str(fila[0] or "").strip().upper(),
            "nombre_proveedora": str(fila[1] or "").strip(),
            "telefono": str(fila[2] or "").strip(),
            "banco": str(fila[3] or "").strip(),
            "numero_cuenta": str(fila[4] or "").strip(),
            "titular_cuenta": str(fila[5] or "").strip(),
            "alias": str(fila[6] or "").strip(),
            "obs_ingreso": str(fila[7] or "").strip(),
            "estado": str(fila[8] or "").strip(),
        })

    if not proveedoras:
        return False, "NO HAY PROVEEDORAS CARGADAS."

    return True, proveedoras


                                            # NUEVA PROVEEDORA
def nueva_proveedora():
    print("\n=== NUEVA PROVEEDORA ===")

    codigo_proveedora = limpiar_texto(input("CODIGO_PROVEEDORA: "))
    nombre_proveedora = limpiar_texto(input("NOMBRE_PROVEEDORA: "))
    telefono = limpiar_texto(input("TELEFONO: "))
    banco = limpiar_texto(input("BANCO: "))
    numero_cuenta = limpiar_texto(input("NUMERO_CUENTA: "))
    titular_cuenta = limpiar_texto(input("TITULAR_CUENTA: "))
    alias = limpiar_texto(input("ALIAS: "))
    obs_ingreso = limpiar_texto(input("OBS_INGRESO: "))

    if not codigo_proveedora or not nombre_proveedora:
        print("ERROR: CODIGO_PROVEEDORA y NOMBRE_PROVEEDORA son obligatorios.")
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_proveedoras = wb["PROVEEDORAS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return

    for fila in ws_proveedoras.iter_rows(min_row=2, values_only=True):
        codigo_existente = fila[0]
        if codigo_existente and str(codigo_existente).strip().upper() == codigo_proveedora:
            print("ERROR: YA EXISTE UNA PROVEEDORA CON ESE CODIGO.")
            return

    ws_proveedoras.append([
        codigo_proveedora,
        nombre_proveedora,
        telefono,
        banco,
        numero_cuenta,
        titular_cuenta,
        alias,
        obs_ingreso,
        "ACTIVA"
    ])

    wb.save(ARCHIVO_EXCEL)

    print("\n=== PROVEEDORA CARGADA CORRECTAMENTE ===")
    print(f"CODIGO_PROVEEDORA: {codigo_proveedora}")
    print(f"NOMBRE_PROVEEDORA: {nombre_proveedora}")
    print(f"ESTADO: ACTIVA")



                                            # VER PROVEEDORA
def ver_proveedora():
    print("\n=== VER PROVEEDORA ===")

    codigo_proveedora = limpiar_texto(input("CODIGO_PROVEEDORA: "))

    if not codigo_proveedora:
        print("ERROR: CODIGO_PROVEEDORA es obligatorio.")
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws_proveedoras = wb["PROVEEDORAS"]
        ws_ingresos = wb["INGRESOS"]
    except PermissionError:
        print("ERROR: Cierra el archivo Excel antes de continuar.")
        return

    proveedora_encontrada = False

    for fila in ws_proveedoras.iter_rows(min_row=2, values_only=True):
        if str(fila[0]).strip().upper() == codigo_proveedora:
            proveedora_encontrada = True

            print("\n=== DATOS DE LA PROVEEDORA ===")
            print(f"CODIGO_PROVEEDORA: {fila[0]}")
            print(f"NOMBRE_PROVEEDORA: {fila[1]}")
            print(f"TELEFONO: {fila[2]}")
            print(f"BANCO: {fila[3]}")
            print(f"NUMERO_CUENTA: {fila[4]}")
            print(f"TITULAR_CUENTA: {fila[5]}")
            print(f"ALIAS: {fila[6]}")
            print(f"OBS_INGRESO: {fila[7]}")
            print(f"ESTADO: {fila[8]}")
            break

    if proveedora_encontrada == False:
        print("ERROR: NO SE ENCONTRO LA PROVEEDORA.")
        return

    total_prendas = 0
    prendas_disponibles = 0
    prendas_vendidas = 0
    prendas_devueltas = 0

    for fila in ws_ingresos.iter_rows(min_row=2, values_only=True):
        codigo_fila = str(fila[1]).strip().upper()
        estado = str(fila[8]).strip().upper()

        if codigo_fila == codigo_proveedora:
            total_prendas += 1

            if estado == "DISPONIBLE":
                prendas_disponibles += 1
            elif estado == "VENDIDO":
                prendas_vendidas += 1
            elif estado in {"DEVUELTO", "PENDIENTE DEVOLUCION"}:
                prendas_devueltas += 1

    print("\n=== RESUMEN DE PRENDAS ===")
    print(f"TOTAL_PRENDAS: {total_prendas}")
    print(f"PRENDAS_DISPONIBLES: {prendas_disponibles}")
    print(f"PRENDAS_VENDIDAS: {prendas_vendidas}")
    print(f"PRENDAS_DEVUELTAS: {prendas_devueltas}")

    print("\n=== DETALLE DE PRENDAS ===")

    for fila in ws_ingresos.iter_rows(min_row=2, values_only=True):
        codigo_fila = str(fila[1]).strip().upper()

        if codigo_fila == codigo_proveedora:
            print(f"""
CODIGO_PRENDA: {fila[2]}
ARTICULO: {fila[3]}
MARCA: {fila[4]}
TALLE: {fila[5]}
COLOR: {fila[6]}
PRECIO: {fila[7]}
ESTADO: {fila[8]}
""")




                                                     # SUBMENU CONSULTAS Y RENDICION
def submenu_consultas_rendicion():
    while True:
        print("\n=== CONSULTAS Y RENDICION ===")
        print("1 - RESUMEN GENERAL DEL MES")
        print("2 - RENDICION DE CUENTAS POR PROVEEDORA")
        print("3 - VENTAS PENDIENTES DE VALIDACION")
        print("4 - VALIDAR VENTA PENDIENTE")
        print("0 - VOLVER")

        opcion = input("ELIGE UNA OPCION: ").strip()

        if opcion == "1":
            resumen_general_mes()

        elif opcion == "2":
            rendicion_cuentas()
       
        elif opcion == "3":
            ventas_pendientes_validacion()

        elif opcion == "4":
            validar_venta_pendiente()

        elif opcion == "0":
            break



                                                        # SUBMENU REMARQUE
def submenu_remarque():
    while True:
        print("\n=== REMARQUE ===")
        print("1 - PRENDAS CON 3 MESES O MAS")
        print("2 - PROPONER REMARQUE")
        print("3 - APROBAR REMARQUE")
        print("0 - VOLVER")

        opcion = input("ELIGE UNA OPCION: ").strip()

        if opcion == "1":
            detectar_prendas_3_meses()

        elif opcion == "2":
            proponer_remarque()

        elif opcion == "3":
            aprobar_remarque()

        elif opcion == "0":
            break

                                                        
                                                         # SUBMENU PROVEEDORAS
def submenu_proveedoras():
    while True:
        print("\n=== PROVEEDORAS ===")
        print("1 - NUEVA PROVEEDORA")
        print("2 - VER PROVEEDORA")
        print("0 - VOLVER")

        opcion = input("ELIGE UNA OPCION: ").strip()

        if opcion == "1":
            nueva_proveedora()

        elif opcion == "2":
            ver_proveedora()
        
        elif opcion == "0":
            break



                                            # MENU PRINCIPAL
def mostrar_menu():
    while True:
        print("1 - CARGAR INGRESO")
        print("2 - CARGAR VENTA")
        print("3 - CONSULTAS Y RENDICION")
        print("4 - EDITAR PRENDA")
        print("5 - REMARQUE")
        print("6 - DEVOLUCION")
        print("7 - PROVEEDORAS")
        print("8 - ELIMINAR INGRESO")
        print("9 - REVERSAR VENTA")
        print("0 - SALIR")

        opcion = input("ELIGE UNA OPCION: ").strip()

        if opcion == "1":
            # PEDIR FECHA UNA SOLA VEZ PARA TODO EL LOTE
            fecha_lote = input("FECHA_INGRESO DEL LOTE (DD-MM-YYYY): ").strip()
            # PEDRI CODIGO_PROVEEDORA UNA SOLA VEZ PARA TODO EL LOTE
            codigo_proveedora = limpiar_texto(input("CODIGO_PROVEEDORA DEL LOTE: ")) 
            if not codigo_proveedora:
                print("ERROR: CODIGO_PROVEEDORA es obligatorio.")
                continue

            # VALIDAR FORMATO DE FECHA DEL LOTE
            try:
                datetime.strptime(fecha_lote, "%d/%m/%Y")
            except ValueError:
                print("ERROR: FECHA_INGRESO invalida. Usa el formato DD/MM/YYYY.")
                continue

            # BUCLE PARA CARGA MULTIPLE DE INGRESOS
            while True:
                cargar_ingreso(fecha_lote, codigo_proveedora)
                otra = input("¿CARGAR OTRO INGRESO? (S/N): ").strip().upper()
                if otra != "S":
                    break

        elif opcion == "2":
            cargar_venta()

        elif opcion == "3":
            submenu_consultas_rendicion()
        
        elif opcion == "4":
            editar_prenda()

        elif opcion == "5":
            submenu_remarque()
            
        elif opcion == "6":
            devolver_prendas_multiples()

        elif opcion == "7":
            submenu_proveedoras()
        
        elif opcion == "8":
            eliminar_ingreso()

        elif opcion == "9":
            reversar_venta()

        elif opcion == "0":
            print("SALIENDO DEL PROGRAMA...")
            break


                                    # INICIO DEL PROGRAMA
if __name__ == "__main__":
    mostrar_menu()
